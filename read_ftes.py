# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx')
ws = wb.worksheets[2]  # FTEs

print('=== ESTRUCTURA COMPLETA DE HOJA FTEs ===')
print()

# Ver fila 2 (encabezados de seccion) - columnas F en adelante
print('FILA 2 (Secciones/Etapas):')
for col in range(6, 42):  # F hasta AP
    val = ws.cell(row=2, column=col).value
    if val:
        letter = get_column_letter(col)
        print(f'  {letter} (col {col}): {val}')

print()
print('FILA 3 (Sub-encabezados F-AN):')
for col in range(6, 42):
    val = ws.cell(row=3, column=col).value
    if val:
        letter = get_column_letter(col)
        print(f'  {letter}: {val}')

print()
print('FILA 4 (Hr/FTE):')
for col in range(6, 42):
    val = ws.cell(row=4, column=col).value
    if val:
        letter = get_column_letter(col)
        print(f'  {letter}: {val}')

# Ver tambien el archivo de referencia (Negociacion Colectiva) para comparar
print()
print('=== COMPARANDO CON ARCHIVO DE REFERENCIA ===')
wb2 = load_workbook(r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Negociacion Colectiva COMPLETO v2.xlsx')
ws2 = wb2.worksheets[2]  # FTEs
print('Secciones en archivo NC:')
for col in range(6, 42):
    val = ws2.cell(row=2, column=col).value
    if val:
        letter = get_column_letter(col)
        print(f'  {letter}: {val}')
