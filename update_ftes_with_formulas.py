# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil
import re

# Copiar archivo
src = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx'
dst = r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_FORMULAS.xlsx'
shutil.copy2(src, dst)

wb = load_workbook(dst)
ws_vs = wb.worksheets[1]  # Value Stream -
ws_ftes = wb.worksheets[2]  # FTEs

print('=== ACTUALIZANDO FTEs CON FORMULAS ===')
print()

# 1. Primero, leer responsables del Value Stream
print('Leyendo Value Stream...')
etapas_responsables = []
for row in range(2, 15):
    n = ws_vs.cell(row=row, column=1).value
    etapa = ws_vs.cell(row=row, column=3).value
    responsable = ws_vs.cell(row=row, column=6).value
    if n is not None and etapa and responsable:
        etapas_responsables.append({'num': int(n), 'etapa': etapa, 'responsable': responsable})

# 2. Normalizar roles
def normalizar_rol(rol):
    rol = rol.strip().replace(')', '').replace('(', '')
    norm = {'eco': 'ECO', 'Eco': 'ECO', 'ECO': 'ECO', 'lpt': 'LPT', 'LPT ': 'LPT', 'LPT': 'LPT',
            'lpm': 'LPM', 'LPM': 'LPM', 'legal': 'Legal', 'Legal ': 'Legal', 'Legal': 'Legal',
            'gt': 'GT', 'GT': 'GT', 'adp': 'ADP', 'ADP': 'ADP', 'Operaciones LPT': 'Operaciones',
            'MEJORA CONTINUA COMPLIANCE': 'Mejora Continua Compliance'}
    return norm.get(rol, rol)

roles_etapas = {}
for er in etapas_responsables:
    for rol in re.split(r'[/,]', er['responsable']):
        rol = normalizar_rol(rol)
        if rol and len(rol) > 1:
            if rol not in roles_etapas:
                roles_etapas[rol] = []
            if er['num'] not in roles_etapas[rol]:
                roles_etapas[rol].append(er['num'])

print(f'Roles encontrados: {list(roles_etapas.keys())}')
print()

# 3. Configurar estructura de columnapa -> (Columna Hr, Columna FTE)
# Según la estructura: Et1=F/G, Et2=H/I, Et3=J/K, etc.
etapa_cols = {
    1: (6, 7),    # F, G
    2: (8, 9),    # H, I
    3: (10, 11),  # J, K
    4: (12, 13),  # L, M
    5: (14, 15),  # N, O
    6: (16, 17),  # P, Q
    7: (18, 19),  # R, S
    8: (20, 21),  # T, U
    9: (22, 23),  # V, W
    10: (24, 25), # X, Y
}

# Columnas de totales
COL_TOTAL_HRS = 26  # Z
COL_TOTAL_FTE = 27  # AA
COL_HRS_SEM = 28    # AB - Celda con valor 44

# Horas estimadas por etapa
horas_etapa = {1:2, 2:3, 3:2, 4:3, 5:3, 6:4, 7:3, 8:2, 9:4, 10:3}

# Info de roles
info_roles = {
    'LPT': ('Líder Personas Tienda', 'Operaciones Tienda', 'LPM'),
    'GT': ('Gerente de Tienda', 'Operaciones Tienda', 'Squad Lead'),
    'Operaciones': ('Operaciones Tienda', 'Operaciones', 'LPM'),
    'ADP': ('Analista de Datos Personas', 'People Analytics', 'Gerente ADP'),
    'LPM': ('Líder Personas Mercado', 'Gestión Personas', 'HRBP'),
    'Mejora Continua Compliance': ('Analista Mejora Continua', 'Compliance', 'Ger. Compliance'),
    'ECO': ('Ejecutivo Clima Org.', 'Gestión Personas', 'LPM'),
    'Legal': ('Abogado Laboral', 'Legal', 'Gerente Legal'),
}

# 4. Des-fusionar celdas y limpiar
print('Des-fusionando celdas...')
merged_ranges = list(ws_ftes.merged_cells.ranges)
for merged_range in merged_ranges:
    try:
        ws_ftes.unmerge_cells(str(merged_range))
        print(f'  Des-fusionado: {merged_range}')
    except:
        pass

print('Limpiando filas de datos...')
for row in range(5, 20):
    for col in range(1, 35):
        try:
            ws_ftes.cell(row=row, column=col).value = None
        except:
            pass

# 5. Escribir celda de referencia para cálculo FTE (44 hrs semanales)
ws_ftes.cell(row=3, column=COL_HRS_SEM, value=44)
print(f'Celda de referencia (44 hrs) en {get_column_letter(COL_HRS_SEM)}3')

# 6. Llenar responsables con fórmulas
print()
print('=== ESCRIBIENDO RESPONSABLES CON FORMULAS ===')

fila = 5
for rol, etapas in sorted(roles_etapas.items()):
    if not rol or len(rol) < 2:
        continue
    
    info = info_roles.get(rol, (rol, 'Por definir', 'Por definir'))
    
    # Datos básicos
    ws_ftes.cell(row=fila, column=1, value=f'RUT-{fila-4}')
    ws_ftes.cell(row=fila, column=2, value=rol)
    ws_ftes.cell(row=fila, column=3, value=info[0])  # Cargo
    ws_ftes.cell(row=fila, column=4, value=info[1])  # Equipo
    ws_ftes.cell(row=fila, column=5, value=info[2])  # Jefatura
    
    # Horas y FTE por etapa
    cols_hr_usadas = []
    cols_fte_usadas = []
    
    for et in etapas:
        if et in etapa_cols:
            col_hr, col_fte = etapa_cols[et]
            hrs = horas_etapa.get(et, 2)
            
            # Escribir horas
            ws_ftes.cell(row=fila, column=col_hr, value=hrs)
            
            # Escribir fórmula FTE: =Hr/$AB$3
            letra_hr = get_column_letter(col_hr)
            formula_fte = f'={letra_hr}{fila}/${get_column_letter(COL_HRS_SEM)}$3'
            ws_ftes.cell(row=fila, column=col_fte, value=formula_fte)
            
            cols_hr_usadas.append(f'{letra_hr}{fila}')
            cols_fte_usadas.append(f'{get_column_letter(col_fte)}{fila}')
    
    # Fórmula TOTAL HRS (suma de todas las horas)
    if cols_hr_usadas:
        formula_total_hrs = '=' + '+'.join(cols_hr_usadas)
        ws_ftes.cell(row=fila, column=COL_TOTAL_HRS, value=formula_total_hrs)
    
    # Fórmula TOTAL FTE (suma de todos los FTE)
    if cols_fte_usadas:
        formula_total_fte = '=' + '+'.join(cols_fte_usadas)
        ws_ftes.cell(row=fila, column=COL_TOTAL_FTE, value=formula_total_fte)
    
    print(f'Fila {fila}: {rol}')
    print(f'         Etapas: {sorted(etapas)} | Hrs: {[horas_etapa.get(e,0) for e in sorted(etapas)]}')
    
    fila += 1

# 7. Actualizar encabezados de totales
try:
    ws_ftes.cell(row=3, column=COL_TOTAL_HRS, value='TOTAL HRS')
except:
    pass
try:
    ws_ftes.cell(row=3, column=COL_TOTAL_FTE, value='TOTAL FTE')
except:
    pass
try:
    ws_ftes.cell(row=4, column=COL_TOTAL_HRS, value='Hrs')
except:
    pass
try:
    ws_ftes.cell(row=4, column=COL_TOTAL_FTE, value='FTE')
except:
    pass

print()
print('Guardando archivo...')
wb.save(dst)

# Copiar de vuelta
shutil.copy2(dst, src)
print('Archivo actualizado en OneDrive!')

print()
print('=== RESUMEN ===')
print(f'Responsables agregados: {fila - 5}')
print(f'Fórmulas FTE: =Hr/${get_column_letter(COL_HRS_SEM)}$3 (divide entre 44 hrs)')
print(f'Total Hrs en columna {get_column_letter(COL_TOTAL_HRS)}')
print(f'Total FTE en columna {get_column_letter(COL_TOTAL_FTE)}')
