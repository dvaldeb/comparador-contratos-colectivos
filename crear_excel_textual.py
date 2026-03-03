# -*- coding: utf-8 -*-
import fitz
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil

contratos_dir = r'c:\Users\dvaldeb\pupy inteeligente\contratos'
output_excel = r'c:\Users\dvaldeb\pupy inteeligente\Clausulas_Contratos_TEXTUAL.xlsx'

# Colores Walmart
WALMART_BLUE = '0053E2'
WALMART_SPARK = 'FFC220'

print('=== EXTRAYENDO CLAUSULAS TEXTUALES ===')
print()

# Temas y palabras clave
temas = {
    'Aniversarios': ['aniversario', 'años de servicio', 'quinquenio'],
    'Excelencia Academica': ['excelencia acad', 'beca', 'becas escolares'],
    'Uniformes': ['uniforme', 'ropa de trabajo', 'vestuario'],
    'Casino': ['casino', 'alimentación', 'colación']
}

# Formatos/marcas
formatos_keywords = {
    'Hiper': ['hiper', 'hipermercado', 'líder'],
    'Express': ['express'],
    'SuperBodega': ['superbodega', 'sba', 'bodega'],
    'Ekono': ['ekono'],
    'Central': ['central', 'oficina'],
    'Todos': ['todos']
}

def extraer_articulo_completo(texto_pagina, keyword):
    """Extrae el artículo completo que contiene la palabra clave"""
    texto_lower = texto_pagina.lower()
    
    if keyword.lower() not in texto_lower:
        return None
    
    # Buscar el inicio del artículo (número seguido de punto o guión)
    # Patrones comunes: "1.-", "Artículo 1", "N° 1", "1)"
    
    idx = texto_lower.find(keyword.lower())
    
    # Buscar hacia atrás el inicio del artículo/sección
    inicio = idx
    for i in range(idx, max(0, idx-500), -1):
        # Buscar patrones de inicio de artículo
        substring = texto_pagina[max(0,i-10):i+5]
        if re.search(r'\d+[.\-\):]\s*[A-ZÁÉÍÓÚ]', substring):
            inicio = i
            break
        if re.search(r'art[\u00edculo\.]*\s*\d+', substring.lower()):
            inicio = max(0, i-10)
            break
    
    # Buscar el final del artículo (siguiente número de artículo o fin de texto)
    fin = len(texto_pagina)
    for i in range(idx + len(keyword), min(len(texto_pagina), idx + 2000)):
        substring = texto_pagina[i:i+15]
        # Buscar siguiente artículo
        if re.search(r'^\s*\d+[.\-\)]\s*[A-ZÁÉÍÓÚ]', substring):
            fin = i
            break
    
    # Extraer el texto
    clausula = texto_pagina[inicio:fin].strip()
    
    # Limpiar un poco pero mantener formato
    clausula = re.sub(r'\n\s*\n', '\n', clausula)  # Reducir líneas vacías múltiples
    
    return clausula

def detectar_formatos(texto):
    """Detecta qué formatos/marcas se mencionan en el texto"""
    texto_lower = texto.lower()
    formatos = []
    
    for formato, keywords in formatos_keywords.items():
        for kw in keywords:
            if kw in texto_lower:
                if formato not in formatos:
                    formatos.append(formato)
                break
    
    return formatos if formatos else ['General']

def buscar_clausulas_textuales(pdf, tema, keywords):
    """Busca y extrae cláusulas textuales completas"""
    resultados = []
    
    for page_num in range(len(pdf)):
        page = pdf[page_num]
        texto = page.get_text()
        
        for kw in keywords:
            if kw.lower() in texto.lower():
                clausula = extraer_articulo_completo(texto, kw)
                
                if clausula and len(clausula) > 50:
                    # Verificar que no sea duplicado
                    es_duplicado = False
                    for r in resultados:
                        if clausula[:100] in r['clausula'] or r['clausula'][:100] in clausula:
                            es_duplicado = True
                            break
                    
                    if not es_duplicado:
                        formatos = detectar_formatos(clausula)
                        resultados.append({
                            'pagina': page_num + 1,
                            'clausula': clausula,
                            'formatos': formatos
                        })
                break
    
    return resultados[:2]  # Max 2 por tema para no saturar

# Procesar contratos
resultados_contratos = {}

for archivo in sorted(os.listdir(contratos_dir)):
    if not archivo.endswith('.pdf'):
        continue
    
    sindicato = archivo.replace('.pdf', '').replace('_', ' ')
    filepath = os.path.join(contratos_dir, archivo)
    
    pdf = fitz.open(filepath)
    
    # Verificar si tiene texto
    texto_test = pdf[0].get_text()
    es_escaneado = len(texto_test.strip()) < 100
    
    print(f'Procesando: {sindicato}')
    
    if es_escaneado:
        print('  [ESCANEADO]')
        resultados_contratos[sindicato] = {
            'formato_pdf': 'Escaneado',
            'total_paginas': len(pdf),
            'temas': {tema: [{'pagina': 'N/A', 'clausula': 'PDF escaneado - Requiere revisión manual del documento original', 'formatos': ['Ver documento']}] for tema in temas.keys()}
        }
    else:
        resultados_contratos[sindicato] = {
            'formato_pdf': 'Texto',
            'total_paginas': len(pdf),
            'temas': {}
        }
        
        for tema, keywords in temas.items():
            clausulas = buscar_clausulas_textuales(pdf, tema, keywords)
            if clausulas:
                resultados_contratos[sindicato]['temas'][tema] = clausulas
                print(f'  {tema}: {len(clausulas)} clausulas encontradas')
            else:
                resultados_contratos[sindicato]['temas'][tema] = [{'pagina': '-', 'clausula': 'No encontrado', 'formatos': ['-']}]
                print(f'  {tema}: No encontrado')
    
    pdf.close()
    print()

# Crear Excel
print('Creando Excel con cláusulas textuales...')

wb = Workbook()
ws = wb.active
ws.title = 'Clausulas Textuales'

# Estilos
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color=WALMART_BLUE, end_color=WALMART_BLUE, fill_type='solid')
subheader_fill = PatternFill(start_color=WALMART_SPARK, end_color=WALMART_SPARK, fill_type='solid')
subheader_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center')

# Título
ws.merge_cells('A1:F1')
ws['A1'] = 'CLÁUSULAS TEXTUALES - CONTRATOS COLECTIVOS'
ws['A1'].font = Font(bold=True, size=16, color=WALMART_BLUE)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:F2')
ws['A2'] = 'Texto extraído directamente de los contratos'
ws['A2'].font = Font(size=11, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Encabezados
headers = ['Sindicato', 'Tema', 'Pág.', 'Formato/Marca', 'CLÁUSULA TEXTUAL', 'Tipo PDF']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# Anchos
ws.column_dimensions['A'].width = 14
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 6
ws.column_dimensions['D'].width = 18
ws.column_dimensions['E'].width = 100  # Columna ancha para cláusula textual
ws.column_dimensions['F'].width = 10

# Datos
row = 5
for sindicato, data in sorted(resultados_contratos.items()):
    first_row = row
    
    for tema, clausulas in data['temas'].items():
        for i, c in enumerate(clausulas):
            # Sindicato
            if row == first_row:
                ws.cell(row=row, column=1, value=sindicato).font = Font(bold=True)
                ws.cell(row=row, column=6, value=data['formato_pdf'])
            
            # Tema
            if i == 0:
                cell_tema = ws.cell(row=row, column=2, value=tema)
                cell_tema.fill = subheader_fill
                cell_tema.font = subheader_font
            
            # Página
            ws.cell(row=row, column=3, value=str(c['pagina']))
            
            # Formato/Marca
            ws.cell(row=row, column=4, value=', '.join(c['formatos']))
            
            # CLÁUSULA TEXTUAL COMPLETA
            ws.cell(row=row, column=5, value=c['clausula'])
            
            # Bordes
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = wrap_align
            
            row += 1
    
    # Merge sindicato
    if row - first_row > 1:
        ws.merge_cells(f'A{first_row}:A{row-1}')
        ws.merge_cells(f'F{first_row}:F{row-1}')
        ws.cell(row=first_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
        ws.cell(row=first_row, column=6).alignment = center_align
    
    row += 1

# Alturas de fila (más grandes para texto completo)
for r in range(5, row):
    ws.row_dimensions[r].height = 150

# Guardar
wb.save(output_excel)
print(f'\nExcel guardado: {output_excel}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Clausulas_Contratos_TEXTUAL.xlsx'
shutil.copy2(output_excel, desktop)
print(f'Copiado a: {desktop}')
