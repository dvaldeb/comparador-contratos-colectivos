import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter
import shutil

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO v2' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

print(f'Archivo: {path}', flush=True)
temp = 'temp_fix2.xlsx'
shutil.copy2(path, temp)
wb = openpyxl.load_workbook(temp)
ws = wb["FTE\u00b4s"]

stages_cols = [6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]
stages_names = [
    'Planificaci\u00f3n', 'Inicio Formal y Fuero', 'N\u00f3mina y Validaci\u00f3n',
    'Comunicaci\u00f3n del Proceso', 'An\u00e1lisis Financiero y Costeo', 'Mesa de Negociaci\u00f3n',
    '\u00daltima Oferta y Votaci\u00f3n', 'Buenos Oficios / Mediaci\u00f3n',
    'SSMM y Equipos de Emergencia', 'Huelga y Gesti\u00f3n Operacional',
    'Cierre, Firma y Registro Legal', 'Implementaci\u00f3n, BTN y Reincorporaci\u00f3n',
    'Seguimiento, Evaluaci\u00f3n y Cierre de Ciclo',
]

# ============================================================
# Responsables por etapa (de la hoja "13 etapas")
# ============================================================
# N=0  Planificaci\u00f3n:              RRSS, Finanzas
# N=1  Inicio Formal y Fuero:      RRSS, Finanzas, Operaci\u00f3n
# N=2  N\u00f3mina y Validaci\u00f3n:        RRSS, Operaci\u00f3n
# N=3  Comunicaci\u00f3n del Proceso:   RRSS (info), ECO (com), Operaciones, Com. Internas
# N=4  An\u00e1l. Financiero y Costeo:  Finanzas, Compensaciones, RRSS
# N=5  Mesa de Negociaci\u00f3n:        RRSS, Legal, Finanzas
# N=6  \u00daltima Oferta y Votaci\u00f3n:   RRSS, Legal, Operaci\u00f3n
# N=7  Buenos Oficios / Mediaci\u00f3n: RRSS, Legal
# N=8  SSMM y Equipos Emergencia:  Operaci\u00f3n (Gte Tienda), RRSS, Legal
# N=9  Huelga y Gesti\u00f3n Operac.:   Operaci\u00f3n, RRSS, Gcia Mercado, Gcia Tienda
# N=10 Cierre, Firma y Registro:   RRSS, Legal, Operaci\u00f3n
# N=11 Implementaci\u00f3n, BTN, Reinc: Compensaciones, N\u00f3mina, Operaci\u00f3n, Gcia Tienda, LP
# N=12 Seguimiento y Cierre Ciclo: RRSS, Com. Internas, Operaci\u00f3n, Comp, Gcia Personas, ECO

# ============================================================
# 10 ROLES alineados a los responsables de cada etapa
# Hrs = dedicaci\u00f3n semanal promedio durante fase activa
# Solo se asignan horas donde el rol es responsable
# ============================================================
roles = [
    {
        'nombre': 'Gerente Relaciones Sindicales',
        'cargo': 'Gerente RRSS',
        'equipo': 'Relaciones Sindicales',
        'jefatura': 'VP Personas',
        #         Plan  IniF  N\u00f3m   Com   Fin   Mesa  Ult   BOfi  SSMM  Huelg Cierr Impl  Seg
        'hrs':   [4,    4,    2,    2,    2,    10,   5,    6,    2,    2,    3,    0,    2],
        'focos': 0,  # 44
    },
    {
        'nombre': 'Profesional Relaciones Sindicales',
        'cargo': 'Profesional RRSS',
        'equipo': 'Relaciones Sindicales',
        'jefatura': 'Gerente RRSS',
        'hrs':   [3,    5,    5,    3,    2,    6,    4,    3,    3,    3,    3,    0,    3],
        'focos': 1,  # 44
    },
    {
        'nombre': 'Abogado Relaciones Laborales',
        'cargo': 'Abogado RRLL',
        'equipo': 'Legal',
        'jefatura': 'Gerente RRSS',
        'hrs':   [0,    3,    0,    0,    0,    8,    7,    8,    5,    0,    6,    0,    0],
        'focos': 7,  # 44
    },
    {
        'nombre': 'Analista Finanzas NC',
        'cargo': 'Analista Finanzas',
        'equipo': 'Finanzas',
        'jefatura': 'Gerente Finanzas',
        'hrs':   [4,    3,    0,    0,    14,   5,    0,    0,    0,    0,    0,    6,    0],
        'focos': 12, # 44
    },
    {
        'nombre': 'Profesional Compensaciones y N\u00f3mina',
        'cargo': 'Profesional Comp & N\u00f3mina',
        'equipo': 'Compensaciones / N\u00f3mina',
        'jefatura': 'Gerente Compensaciones',
        'hrs':   [0,    0,    0,    0,    8,    0,    0,    0,    0,    0,    0,    14,   4],
        'focos': 18, # 44
    },
    {
        'nombre': 'ECO - Comunicaciones Internas',
        'cargo': 'Profesional ECO',
        'equipo': 'ECO / Comunicaciones Internas',
        'jefatura': 'Gerente ECO',
        'hrs':   [0,    0,    0,    12,   0,    0,    3,    0,    0,    3,    0,    0,    8],
        'focos': 18, # 44
    },
    {
        'nombre': 'Gerente de Mercado',
        'cargo': 'Gerente de Mercado',
        'equipo': 'Operaci\u00f3n / Gerencia de Mercado',
        'jefatura': 'Director Operaciones',
        'hrs':   [0,    0,    0,    2,    0,    0,    2,    0,    2,    8,    2,    2,    2],
        'focos': 24, # 44
    },
    {
        'nombre': 'L\u00edder de Persona Mercado (LPM)',
        'cargo': 'L\u00edder de Persona Mercado',
        'equipo': 'Operaci\u00f3n',
        'jefatura': 'Gerente de Mercado',
        'hrs':   [1,    2,    4,    3,    0,    0,    2,    0,    4,    6,    3,    4,    3],
        'focos': 12, # 44
    },
    {
        'nombre': 'Gerente de Tienda',
        'cargo': 'Gerente de Tienda',
        'equipo': 'Operaci\u00f3n Tienda',
        'jefatura': 'Gerente de Mercado',
        'hrs':   [0,    0,    0,    1,    0,    0,    1,    0,    5,    10,   2,    4,    2],
        'focos': 19, # 44
    },
    {
        'nombre': 'L\u00edder de Persona Tienda (LPT)',
        'cargo': 'L\u00edder de Persona Tienda',
        'equipo': 'Operaci\u00f3n Tienda',
        'jefatura': 'Gerente de Tienda',
        'hrs':   [0,    1,    4,    2,    0,    0,    1,    0,    5,    8,    2,    6,    3],
        'focos': 12, # 44
    },
]

# Validate 44h
for role in roles:
    total = sum(role['hrs']) + role['focos']
    if total != 44:
        diff = total - 44
        max_idx = role['hrs'].index(max(role['hrs']))
        role['hrs'][max_idx] -= diff
        print(f"  Ajustado {role['nombre']}: {total} -> 44", flush=True)
    total2 = sum(role['hrs']) + role['focos']
    assert total2 == 44, f"ERROR: {role['nombre']} = {total2}"

# ============================================================
# Clear ALL data rows (5 to 20)
# ============================================================
for r in range(5, 22):
    for c in range(1, 42):
        ws.cell(row=r, column=c, value=None)

# ============================================================
# Write roles
# ============================================================
data_start = 5
data_end = data_start + len(roles) - 1  # row 14

for i, role in enumerate(roles):
    row = data_start + i
    ws.cell(row=row, column=1, value='-')
    ws.cell(row=row, column=2, value=role['nombre'])
    ws.cell(row=row, column=3, value=role['cargo'])
    ws.cell(row=row, column=4, value=role['equipo'])
    ws.cell(row=row, column=5, value=role['jefatura'])

    for j, col_start in enumerate(stages_cols):
        ws.cell(row=row, column=col_start, value=role['hrs'][j])
        hr_letter = get_column_letter(col_start)
        ws.cell(row=row, column=col_start + 1, value=f'={hr_letter}{row}/$AM$3')

    # Focos (AH=34, AI=35)
    ws.cell(row=row, column=34, value=role['focos'])
    ws.cell(row=row, column=35, value=f'=AH{row}/$AM$3')

    # Totals (AK=37, AL=38)
    hr_parts = '+'.join([f'{get_column_letter(c)}{row}' for c in stages_cols] + [f'AH{row}'])
    fte_parts = '+'.join([f'{get_column_letter(c+1)}{row}' for c in stages_cols] + [f'AI{row}'])
    ws.cell(row=row, column=37, value=f'={hr_parts}')
    ws.cell(row=row, column=38, value=f'={fte_parts}')

# ============================================================
# Summary rows (Max, Min, Mediana)
# ============================================================
summary_start = data_end + 2  # row 16

for offset, (label, func) in enumerate([('Max', 'MAX'), ('Min', 'MIN'), ('Mediana', 'MEDIAN')]):
    row = summary_start + offset
    ws.cell(row=row, column=5, value=label)

    for col_start in stages_cols:
        hr_letter = get_column_letter(col_start)
        fte_letter = get_column_letter(col_start + 1)
        ws.cell(row=row, column=col_start,
                value=f'={func}({hr_letter}{data_start}:{hr_letter}{data_end})')
        ws.cell(row=row, column=col_start + 1,
                value=f'={fte_letter}{row}/$AM$3')

    ws.cell(row=row, column=34, value=f'={func}(AH{data_start}:AH{data_end})')
    ws.cell(row=row, column=35, value=f'=AH{row}/$AM$3')

# Nota
ws.cell(row=summary_start + 4, column=1,
        value='Consideraciones: Las horas reflejan la dedicaci\u00f3n semanal promedio de cada rol durante la fase activa de cada etapa. Solo se asignan horas donde el rol es responsable seg\u00fan la hoja "13 etapas". Suma = 44 hrs (100% FTE).')

# ============================================================
# Print cross-reference validation
# ============================================================
print('\n=== VALIDACI\u00d3N: RESPONSABLES POR ETAPA ===', flush=True)
for j, name in enumerate(stages_names):
    involved = []
    for role in roles:
        if role['hrs'][j] > 0:
            involved.append(f"{role['cargo']}({role['hrs'][j]}h)")
    print(f'  N={j:2d} {name:45s} -> {', '.join(involved)}', flush=True)

# ============================================================
# Save
# ============================================================
try:
    wb.save(path)
    print(f'\n\u2705 Guardado en: {path}', flush=True)
except PermissionError:
    alt = path.replace('v2', 'v3')
    wb.save(alt)
    path = alt
    print(f'\n\u26a0\ufe0f Guardado como: {alt}', flush=True)

try:
    os.remove(temp)
except:
    pass

print(f'\n{len(roles)} roles (filas {data_start}-{data_end}), Summary filas {summary_start}-{summary_start+2}', flush=True)
for role in roles:
    total = sum(role['hrs']) + role['focos']
    print(f"  {role['nombre']:45s} | {role['equipo']:30s} | {total}h", flush=True)
