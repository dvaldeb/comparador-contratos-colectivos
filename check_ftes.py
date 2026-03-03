# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook(r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx')
ws = wb.worksheets[2]  # FTEs

print('=== VERIFICANDO HOJA FTEs ===')
print(f'Nombre hoja: {ws.title}')
print()

print('Filas 5-12:')
for row in range(5, 13):
    vals = [ws.cell(row=row, column=c).value for c in range(1, 10)]
    print(f'Fila {row}: {vals}')
