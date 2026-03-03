import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter
import shutil

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO v2' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

print(f'Archivo: {path}', flush=True)
temp = 'temp_v2c.xlsx'
shutil.copy2(path, temp)
wb = openpyxl.load_workbook(temp)
ws = wb["FTE\u00b4s"]

# Stage columns (13 stages: col 6 to 31, each 2 cols)
stages_cols = [
    6,   # 0  Planificaci\u00f3n
    8,   # 1  Inicio Formal y Fuero
    10,  # 2  N\u00f3mina y Validaci\u00f3n
    12,  # 3  Comunicaci\u00f3n del Proceso
    14,  # 4  An\u00e1lisis Financiero y Costeo
    16,  # 5  Mesa de Negociaci\u00f3n
    18,  # 6  \u00daltima Oferta y Votaci\u00f3n
    20,  # 7  Buenos Oficios / Mediaci\u00f3n
    22,  # 8  SSMM y Equipos de Emergencia
    24,  # 9  Huelga y Gesti\u00f3n Operacional
    26,  # 10 Cierre, Firma y Registro Legal
    28,  # 11 Implementaci\u00f3n, BTN y Reincorporaci\u00f3n
    30,  # 12 Seguimiento, Evaluaci\u00f3n y Cierre
]

# ============================================================
# 8 ROLES del proceso de NC
# Hrs = horas semanales promedio durante fase activa de cada etapa
# ============================================================
roles = [
    {
        'nombre': 'Gerente Relaciones Sindicales',
        'cargo': 'Gerente RRSS',
        'equipo': 'Relaciones Sindicales',
        'jefatura': 'VP Personas',
        #         Plan  IniF  N\u00f3m   Com   Fin   Mesa  Ult   BOfi  SSMM  Huelg Cierr Impl  Seg
        'hrs':   [4,    4,    2,    2,    3,    10,   5,    5,    2,    2,    3,    1,    1],
        'focos': 0,
    },
    {
        'nombre': 'Profesional Relaciones Sindicales',
        'cargo': 'Profesional RRSS',
        'equipo': 'Relaciones Sindicales',
        'jefatura': 'Gerente RRSS',
        'hrs':   [3,    5,    5,    2,    2,    6,    4,    3,    3,    3,    3,    2,    2],
        'focos': 1,
    },
    {
        'nombre': 'Abogado Relaciones Laborales',
        'cargo': 'Abogado RRLL',
        'equipo': 'Relaciones Sindicales / Legal',
        'jefatura': 'Gerente RRSS',
        'hrs':   [2,    4,    2,    1,    1,    6,    6,    6,    4,    2,    5,    1,    1],
        'focos': 3,
    },
    {
        'nombre': 'Analista Finanzas NC (Cami / Juan Millanao)',
        'cargo': 'Analista Finanzas',
        'equipo': 'Finanzas',
        'jefatura': 'Gerente Finanzas',
        'hrs':   [3,    2,    3,    1,    12,   4,    3,    0,    0,    0,    2,    6,    1],
        'focos': 7,
    },
    {
        'nombre': 'Profesional Compensaciones',
        'cargo': 'Profesional Compensaciones',
        'equipo': 'Compensaciones',
        'jefatura': 'Gerente Compensaciones',
        'hrs':   [2,    1,    1,    0,    8,    3,    2,    0,    0,    0,    2,    12,   2],
        'focos': 11,
    },
    {
        'nombre': 'ECO - Comunicaciones Operaci\u00f3n',
        'cargo': 'Profesional ECO',
        'equipo': 'ECO / Comunicaciones Internas',
        'jefatura': 'Gerente ECO',
        'hrs':   [1,    2,    1,    10,   1,    2,    3,    1,    2,    3,    2,    2,    6],
        'focos': 8,
    },
    {
        'nombre': 'L\u00edder de Persona Mercado (LPM)',
        'cargo': 'L\u00edder de Persona Mercado',
        'equipo': 'Operaci\u00f3n',
        'jefatura': 'Gerente de Mercado',
        'hrs':   [1,    2,    3,    3,    1,    1,    2,    1,    4,    6,    2,    4,    3],
        'focos': 11,
    },
    {
        'nombre': 'L\u00edder de Persona Tienda (LPT)',
        'cargo': 'L\u00edder de Persona Tienda',
        'equipo': 'Operaci\u00f3n Tienda',
        'jefatura': 'Gerente de Tienda',
        'hrs':   [0,    1,    3,    2,    0,    0,    1,    0,    6,    10,   2,    5,    3],
        'focos': 11,
    },
]

# Validate all sum to 44
for role in roles:
    total = sum(role['hrs']) + role['focos']
    if total != 44:
        diff = total - 44
        max_idx = role['hrs'].index(max(role['hrs']))
        role['hrs'][max_idx] -= diff
        print(f"  Ajustado {role['nombre']}: {total} -> 44", flush=True)
    total = sum(role['hrs']) + role['focos']
    assert total == 44, f"ERROR: {role['nombre']} = {total}"

# ============================================================
# Clear data rows 5-16
# ============================================================
for r in range(5, 20):
    for c in range(1, 42):
        ws.cell(row=r, column=c, value=None)

# ============================================================
# Write 8 roles (rows 5-12)
# ============================================================
data_start = 5
data_end = data_start + len(roles) - 1  # row 12

for i, role in enumerate(roles):
    row = data_start + i
    ws.cell(row=row, column=1, value='-')                # A: Rut
    ws.cell(row=row, column=2, value=role['nombre'])     # B: Nombre
    ws.cell(row=row, column=3, value=role['cargo'])      # C: Cargo
    ws.cell(row=row, column=4, value=role['equipo'])     # D: Equipo
    ws.cell(row=row, column=5, value=role['jefatura'])   # E: Jefatura

    # Hours for each of the 13 stages
    for j, col_start in enumerate(stages_cols):
        ws.cell(row=row, column=col_start, value=role['hrs'][j])  # Hr
        hr_letter = get_column_letter(col_start)
        ws.cell(row=row, column=col_start + 1, value=f'={hr_letter}{row}/$AM$3')  # FTE

    # Focos Especificos (AH=34, AI=35)
    ws.cell(row=row, column=34, value=role['focos'])
    ws.cell(row=row, column=35, value=f'=AH{row}/$AM$3')

    # Total Hrs (AK=37)
    hr_parts = '+'.join([f'{get_column_letter(c)}{row}' for c in stages_cols] + [f'AH{row}'])
    ws.cell(row=row, column=37, value=f'={hr_parts}')

    # Total FTE (AL=38)
    fte_parts = '+'.join([f'{get_column_letter(c+1)}{row}' for c in stages_cols] + [f'AI{row}'])
    ws.cell(row=row, column=38, value=f'={fte_parts}')

# ============================================================
# Summary rows: Max, Min, Mediana (rows 14-16)
# ============================================================
summary_start = data_end + 2  # row 14

for offset, (label, func) in enumerate([('Max', 'MAX'), ('Min', 'MIN'), ('Mediana', 'MEDIAN')]):
    row = summary_start + offset
    ws.cell(row=row, column=5, value=label)

    for col_start in stages_cols:
        hr_letter = get_column_letter(col_start)
        fte_letter = get_column_letter(col_start + 1)
        ws.cell(row=row, column=col_start, value=f'={func}({hr_letter}{data_start}:{hr_letter}{data_end})')
        ws.cell(row=row, column=col_start + 1, value=f'={fte_letter}{row}/$AM$3')

    # Focos
    ws.cell(row=row, column=34, value=f'={func}(AH{data_start}:AH{data_end})')
    ws.cell(row=row, column=35, value=f'=AH{row}/$AM$3')

# Consideraciones (row 18)
ws.cell(row=summary_start + 4, column=1,
        value='Consideraciones: Las horas reflejan la dedicaci\u00f3n semanal promedio de cada rol durante la fase activa de cada etapa del proceso de NC. La suma no puede exceder 44 hrs (jornada completa).')

# ============================================================
# Save
# ============================================================
try:
    wb.save(path)
    print(f'\n\u2705 Guardado en: {path}', flush=True)
except PermissionError:
    alt = path.replace('v2', 'v3')
    wb.save(alt)
    path = alt
    print(f'\n\u26a0\ufe0f Guardado como: {alt}', flush=True)

try:
    os.remove(temp)
except:
    pass

print(f'\n8 roles escritos en filas {data_start}-{data_end}', flush=True)
print(f'Summary en filas {summary_start}-{summary_start+2}', flush=True)
print('\nDetalle:', flush=True)
for role in roles:
    total = sum(role['hrs']) + role['focos']
    print(f"  {role['nombre']:45s} | {role['equipo']:30s} | {total}h", flush=True)
