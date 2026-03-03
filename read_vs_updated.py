# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import shutil

# Leer archivo
wb = load_workbook(r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx')

# La hoja ahora se llama "Value Stream -"
ws = wb.worksheets[1]  # Segunda hoja (Value Stream)

print('=== DATOS ACTUALES DEL VALUE STREAM ===')
print(f'Nombre de la hoja: {ws.title}')
print()

print('Encabezados:', [c.value for c in ws[1]])
print()

print('ETAPAS Y RESPONSABLES:')
etapas_data = []
for row in range(2, 30):  # Revisar varias filas
    n = ws.cell(row=row, column=1).value
    valor = ws.cell(row=row, column=2).value
    etapa = ws.cell(row=row, column=3).value
    proceso = ws.cell(row=row, column=4).value
    descripcion = ws.cell(row=row, column=5).value
    responsable = ws.cell(row=row, column=6).value
    desc_proceso = ws.cell(row=row, column=7).value
    
    if n is not None and (etapa or responsable):
        print(f'---')
        print(f'Fila {row} | N={n}')
        print(f'  Etapa: {etapa}')
        print(f'  VALOR: {valor}')
        print(f'  Proceso: {proceso}')
        print(f'  Responsable: {responsable}')
        
        etapas_data.append({
            'num': n,
            'etapa': etapa,
            'responsable': responsable,
            'valor': valor
        })

print()
print(f'Total etapas encontradas: {len(etapas_data)}')
