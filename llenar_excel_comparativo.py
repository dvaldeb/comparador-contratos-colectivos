# -*- coding: utf-8 -*-
import fitz
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment
import shutil

print('=== EXTRAYENDO CLAUSULAS Y LLENANDO EXCEL ===')
print()

# Configuración de temas
temas = {
    'Excelencia Académica': {
        'keywords': ['excelencia acad', 'beca', 'becas', 'escolaridad', 'premio mejor'],
        'fila_excel': 2
    },
    'Aniversarios': {
        'keywords': ['aniversario', 'años de servicio', 'celebrarán su aniversario'],
        'fila_excel': 3
    },
    'Casinos': {
        'keywords': ['casino', 'alimentación', 'colación', 'bono de alimentación'],
        'fila_excel': 4
    },
    'Uniformes': {
        'keywords': ['uniforme', 'ropa de trabajo', 'vestuario', 'vestimenta'],
        'fila_excel': 5
    }
}

# PDFs a procesar con su columna en el Excel
pdfs_config = {
    'SIL_2025.pdf': {
        'path': r'c:\Users\dvaldeb\pupy inteeligente\contratos\SIL_2025.pdf',
        'columna': 4,  # Columna D (SIL)
        'nombre': 'SIL',
        'version': 'CC SIL 2025-2027 Firma 02.07.2025'
    },
    'FED_WM_2025.pdf': {
        'path': r'c:\Users\dvaldeb\pupy inteeligente\contratos\FED_WM_2025.pdf',
        'columna': 3,  # Columna C (FNW)
        'nombre': 'FED WM / FNW',
        'version': 'CC FED WM 15.12.2025 (Versión Definitiva)'
    }
}

def extraer_clausula_completa(pdf, keywords):
    """Busca y extrae la cláusula completa"""
    for page_num in range(len(pdf)):
        page = pdf[page_num]
        texto = page.get_text()
        texto_lower = texto.lower()
        
        for kw in keywords:
            if kw.lower() in texto_lower:
                # Encontrar el contexto completo
                idx = texto_lower.find(kw.lower())
                
                # Buscar inicio del artículo/sección
                inicio = max(0, idx - 200)
                for i in range(idx, max(0, idx-200), -1):
                    if re.search(r'\d+[.\-\)]\s*[A-Z]', texto[max(0,i-5):i+5]):
                        inicio = max(0, i-5)
                        break
                
                # Buscar fin del artículo
                fin = min(len(texto), idx + 1200)
                for i in range(idx + 100, min(len(texto), idx + 1200)):
                    if re.search(r'^\s*\d+[.\-\)]\s*[A-Z]', texto[i:i+15]):
                        fin = i
                        break
                
                clausula = texto[inicio:fin].strip()
                clausula = re.sub(r'\n\s*\n+', '\n', clausula)
                
                if len(clausula) > 50:
                    return {
                        'pagina': page_num + 1,
                        'clausula': clausula
                    }
    
    return None

# Cargar Excel
excel_path = r'c:\Users\dvaldeb\pupy inteeligente\Cuadro_Comparativo.xlsx'
wb = load_workbook(excel_path)
ws = wb['Hoja1']

# Procesar cada PDF
resultados = {}

for pdf_name, config in pdfs_config.items():
    print(f'Procesando: {config["nombre"]}')
    pdf = fitz.open(config['path'])
    
    resultados[config['nombre']] = {}
    
    for tema, tema_config in temas.items():
        resultado = extraer_clausula_completa(pdf, tema_config['keywords'])
        
        if resultado:
            resultados[config['nombre']][tema] = resultado
            print(f'  {tema}: Pág {resultado["pagina"]}')
            
            # Escribir en Excel
            fila = tema_config['fila_excel']
            col_clausula = config['columna']
            
            # Escribir cláusula
            ws.cell(row=fila, column=col_clausula, value=resultado['clausula'])
            ws.cell(row=fila, column=col_clausula).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Escribir página (columna H = 8)
            pagina_actual = ws.cell(row=fila, column=8).value or ''
            nueva_pagina = f"{config['nombre']}: Pág {resultado['pagina']}"
            if pagina_actual:
                ws.cell(row=fila, column=8, value=f"{pagina_actual}\n{nueva_pagina}")
            else:
                ws.cell(row=fila, column=8, value=nueva_pagina)
            
            # Escribir versión contrato (columna I = 9)
            version_actual = ws.cell(row=fila, column=9).value or ''
            if config['version'] not in str(version_actual):
                if version_actual:
                    ws.cell(row=fila, column=9, value=f"{version_actual}\n{config['version']}")
                else:
                    ws.cell(row=fila, column=9, value=config['version'])
        else:
            print(f'  {tema}: No encontrado')
    
    pdf.close()
    print()

# Ajustar alturas de fila
for row in range(2, 7):
    ws.row_dimensions[row].height = 150

# Guardar Excel
wb.save(excel_path)
print(f'Excel actualizado: {excel_path}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Cuadro comparativo 4 Beneficios_ACTUALIZADO.xlsx'
shutil.copy2(excel_path, desktop)
print(f'Copiado a: {desktop}')

# Copiar también a Downloads (reemplazar original)
shutil.copy2(excel_path, r'c:\Users\dvaldeb\Downloads\Cuadro comparativo 4 Beneficios.xlsx')
print('Original actualizado en Downloads')
