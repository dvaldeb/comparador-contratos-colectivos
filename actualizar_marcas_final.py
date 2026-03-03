# -*- coding: utf-8 -*-
import fitz
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import shutil

print('=== ACTUALIZANDO EXCEL CON MARCAS CORRECTAS ===')
print()

# Marcas por contrato (extraídas de los encabezados)
MARCAS_CONTRATOS = {
    'SIL': 'Hiper, Express, Ekono, Walmart Mayorista',
    'FED WM / FNW': 'Hiper, Express, Ekono, Acuenta, SBA'
}

# Configuración de temas
temas = {
    'Excelencia Académica': {
        'keywords': ['excelencia acad', 'beca', 'becas', 'escolaridad', 'premio mejor'],
        'fila_excel': 2
    },
    'Aniversarios': {
        'keywords': ['aniversario', 'años de servicio', 'celebrarán su aniversario'],
        'fila_excel': 3
    },
    'Casinos': {
        'keywords': ['casino', 'alimentación', 'colación', 'bono de alimentación'],
        'fila_excel': 4
    },
    'Uniformes': {
        'keywords': ['uniforme', 'ropa de trabajo', 'vestuario', 'vestimenta'],
        'fila_excel': 5
    }
}

# PDFs a procesar
pdfs_config = {
    'SIL_2025.pdf': {
        'path': r'c:\Users\dvaldeb\pupy inteeligente\contratos\SIL_2025.pdf',
        'columna': 4,  # Columna D (SIL)
        'nombre': 'SIL',
        'version': 'CC SIL 2025-2027 Firma 02.07.2025',
        'marcas': 'Hiper, Express, Ekono, Walmart Mayorista'
    },
    'FED_WM_2025.pdf': {
        'path': r'c:\Users\dvaldeb\pupy inteeligente\contratos\FED_WM_2025.pdf',
        'columna': 3,  # Columna C (FNW)
        'nombre': 'FED WM / FNW',
        'version': 'CC FED WM 15.12.2025 (Versión Definitiva)',
        'marcas': 'Hiper, Express, Ekono, Acuenta, SBA'
    }
}

def extraer_clausula_completa(pdf, keywords):
    """Busca y extrae la cláusula completa"""
    for page_num in range(len(pdf)):
        page = pdf[page_num]
        texto = page.get_text()
        texto_lower = texto.lower()
        
        for kw in keywords:
            if kw.lower() in texto_lower:
                idx = texto_lower.find(kw.lower())
                
                # Buscar inicio del artículo
                inicio = max(0, idx - 200)
                for i in range(idx, max(0, idx-200), -1):
                    if re.search(r'\d+[.\-\)]\s*[A-Z]', texto[max(0,i-5):i+5]):
                        inicio = max(0, i-5)
                        break
                
                # Buscar fin del artículo (más amplio)
                fin = min(len(texto), idx + 2000)
                for i in range(idx + 100, min(len(texto), idx + 2000)):
                    if re.search(r'^\s*\d+[.\-\)]\s*[A-Z]', texto[i:i+15]):
                        fin = i
                        break
                
                clausula = texto[inicio:fin].strip()
                clausula = re.sub(r'\n\s*\n+', '\n', clausula)
                
                if len(clausula) > 50:
                    return {
                        'pagina': page_num + 1,
                        'clausula': clausula
                    }
    
    return None

# Cargar Excel original de Downloads
excel_original = r'c:\Users\dvaldeb\Downloads\Cuadro comparativo 4 Beneficios.xlsx'
wb = load_workbook(excel_original)
ws = wb['Hoja1']

# Limpiar columnas que vamos a llenar
for fila in range(2, 6):
    ws.cell(row=fila, column=2, value='')  # Marca
    ws.cell(row=fila, column=8, value='')  # Página
    ws.cell(row=fila, column=9, value='')  # Versión

# Procesar cada PDF
for pdf_name, config in pdfs_config.items():
    print(f'Procesando: {config["nombre"]}')
    print(f'  Marcas: {config["marcas"]}')
    pdf = fitz.open(config['path'])
    
    for tema, tema_config in temas.items():
        resultado = extraer_clausula_completa(pdf, tema_config['keywords'])
        
        if resultado:
            fila = tema_config['fila_excel']
            col_clausula = config['columna']
            
            print(f'  {tema}: Pág {resultado["pagina"]}')
            
            # 1. Escribir CLÁUSULA TEXTUAL (columna del sindicato)
            ws.cell(row=fila, column=col_clausula, value=resultado['clausula'])
            ws.cell(row=fila, column=col_clausula).alignment = Alignment(wrap_text=True, vertical='top')
            
            # 2. Escribir MARCA (columna B = 2)
            marca_actual = ws.cell(row=fila, column=2).value or ''
            nueva_marca = f"{config['nombre']}: {config['marcas']}"
            if nueva_marca not in str(marca_actual):
                if marca_actual:
                    ws.cell(row=fila, column=2, value=f"{marca_actual}\n\n{nueva_marca}")
                else:
                    ws.cell(row=fila, column=2, value=nueva_marca)
            ws.cell(row=fila, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            
            # 3. Escribir PÁGINA (columna H = 8)
            pagina_actual = ws.cell(row=fila, column=8).value or ''
            nueva_pagina = f"{config['nombre']}: Pág {resultado['pagina']}"
            if nueva_pagina not in str(pagina_actual):
                if pagina_actual:
                    ws.cell(row=fila, column=8, value=f"{pagina_actual}\n{nueva_pagina}")
                else:
                    ws.cell(row=fila, column=8, value=nueva_pagina)
            ws.cell(row=fila, column=8).alignment = Alignment(wrap_text=True, vertical='top')
            
            # 4. Escribir VERSIÓN CONTRATO (columna I = 9)
            version_actual = ws.cell(row=fila, column=9).value or ''
            if config['version'] not in str(version_actual):
                if version_actual:
                    ws.cell(row=fila, column=9, value=f"{version_actual}\n{config['version']}")
                else:
                    ws.cell(row=fila, column=9, value=config['version'])
            ws.cell(row=fila, column=9).alignment = Alignment(wrap_text=True, vertical='top')
        else:
            print(f'  {tema}: No encontrado')
    
    pdf.close()
    print()

# Ajustar alturas y anchos
for row in range(2, 7):
    ws.row_dimensions[row].height = 200

ws.column_dimensions['B'].width = 40  # Marca
ws.column_dimensions['C'].width = 60  # FNW
ws.column_dimensions['D'].width = 60  # SIL
ws.column_dimensions['H'].width = 20  # Página
ws.column_dimensions['I'].width = 35  # Versión

# Guardar
wb.save(excel_original)
print(f'Excel actualizado: {excel_original}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Cuadro comparativo 4 Beneficios_FINAL.xlsx'
shutil.copy2(excel_original, desktop)
print(f'Copiado a: {desktop}')
