import sys, os, glob
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# Find the exact file
desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
for f in os.listdir(desktop):
    if 'Copia' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        print(f'Found: {path}', flush=True)
        print(f'Exists: {os.path.exists(path)}', flush=True)
        
        wb = openpyxl.load_workbook(path)
        print(f'Sheets: {wb.sheetnames}', flush=True)
        
        for sn in wb.sheetnames:
            ws = wb[sn]
            print(f'\n======= Sheet: "{sn}" =======', flush=True)
            print(f'Max row: {ws.max_row}, Max col: {ws.max_column}', flush=True)
            for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 40), values_only=False):
                vals = []
                for cell in row:
                    if cell.value is not None:
                        vals.append((cell.column_letter, cell.row, repr(cell.value)[:80]))
                if vals:
                    print(f'  Row {row[0].row}: {vals}', flush=True)
        break
