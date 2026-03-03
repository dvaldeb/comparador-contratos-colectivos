import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

path = 'temp_mapeo.xlsx'
wb = openpyxl.load_workbook(path, data_only=True)
print(f'Sheets: {wb.sheetnames}', flush=True)

for sn in wb.sheetnames:
    ws = wb[sn]
    print(f'\n========== {sn} ========== rows={ws.max_row} cols={ws.max_column}', flush=True)
    print(f'Merged cells: {[str(m) for m in ws.merged_cells.ranges][:20]}', flush=True)
    for r in range(1, min(ws.max_row + 1, 60)):
        vals = []
        for c in range(1, min(ws.max_column + 1, 30)):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                vals.append(f'R{r}C{c}={repr(v)[:100]}')
        if vals:
            print('  ' + ' | '.join(vals), flush=True)
