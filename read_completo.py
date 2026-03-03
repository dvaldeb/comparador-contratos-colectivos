import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

print(f'Archivo: {path}', flush=True)
wb = openpyxl.load_workbook(path)
ws = wb['Value Stream - AT']
print(f'Max row: {ws.max_row}, Max col: {ws.max_column}', flush=True)
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=False):
    a = row[0].value  # N
    c = row[2].value if len(row) > 2 else None  # Etapa
    d = row[3].value if len(row) > 3 else None  # Proceso
    if a is not None or c is not None:
        print(f'  Row {row[0].row}: N={a} | Etapa={c} | Proceso={str(d)[:60] if d else ""}', flush=True)
