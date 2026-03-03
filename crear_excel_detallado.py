# -*- coding: utf-8 -*-
import fitz
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil

contratos_dir = r'c:\Users\dvaldeb\pupy inteeligente\contratos'
output_excel = r'c:\Users\dvaldeb\pupy inteeligente\Clausulas_Contratos_Detallado.xlsx'

# Colores Walmart
WALMART_BLUE = '0053E2'
WALMART_SPARK = 'FFC220'
LIGHT_BLUE = 'E6F0FF'
GREEN = '2A8703'
GRAY = 'F5F5F5'

print('=== EXTRAYENDO CLAUSULAS DETALLADAS ===')
print()

# Temas y palabras clave
temas = {
    'Aniversarios': ['aniversario', 'años de servicio', 'quinquenio', 'antiguedad', 'años de servicios'],
    'Excelencia Academica': ['excelencia acad', 'beca', 'becas', 'escolaridad', 'rendimiento escolar', 'premio mejor'],
    'Uniformes': ['uniforme', 'uniformes', 'vestuario', 'vestimenta', 'ropa de trabajo'],
    'Casino': ['casino', 'alimentación', 'colación', 'bono alimentaci', 'comedor']
}

# Formatos/marcas a buscar
formatos_keywords = {
    'Hiper': ['hiper', 'hipermercado', 'líder'],
    'Express': ['express', 'lider express'],
    'SuperBodega': ['superbodega', 'sba', 'super bodega', 'bodega'],
    'Ekono': ['ekono', 'ekóno'],
    'Central': ['central', 'oficina central', 'casa matriz'],
    'Mayorista': ['mayorista', 'acuenta'],
    'CD': ['centro de distribución', 'cd ', 'bodega central'],
    'Todos': ['todos los formatos', 'todas las tiendas', 'todos los trabajadores']
}

def extraer_clausula_detallada(pdf, page_num, keywords):
    """Extrae el texto de la cláusula y detecta formatos"""
    page = pdf[page_num]
    texto = page.get_text()
    texto_lower = texto.lower()
    
    # Buscar el párrafo que contiene la palabra clave
    for kw in keywords:
        if kw.lower() in texto_lower:
            # Encontrar el contexto
            idx = texto_lower.find(kw.lower())
            # Extraer desde 100 caracteres antes hasta 500 después
            inicio = max(0, idx - 100)
            fin = min(len(texto), idx + 500)
            contexto = texto[inicio:fin]
            
            # Limpiar
            contexto = ' '.join(contexto.split())
            
            # Detectar formatos mencionados
            formatos_encontrados = []
            for formato, fkws in formatos_keywords.items():
                for fkw in fkws:
                    if fkw.lower() in texto_lower:
                        if formato not in formatos_encontrados:
                            formatos_encontrados.append(formato)
            
            if not formatos_encontrados:
                formatos_encontrados = ['General']
            
            return contexto, formatos_encontrados
    
    return None, ['General']

def buscar_clausulas_por_tema(pdf, tema, keywords):
    """Busca todas las páginas donde aparece el tema"""
    resultados = []
    
    for page_num in range(len(pdf)):
        page = pdf[page_num]
        texto = page.get_text().lower()
        
        for kw in keywords:
            if kw.lower() in texto:
                clausula, formatos = extraer_clausula_detallada(pdf, page_num, keywords)
                if clausula and len(clausula) > 50:
                    resultados.append({
                        'pagina': page_num + 1,
                        'clausula': clausula[:600],
                        'formatos': formatos
                    })
                break
    
    # Eliminar duplicados y limitar
    seen = set()
    unique = []
    for r in resultados:
        key = r['pagina']
        if key not in seen:
            seen.add(key)
            unique.append(r)
    
    return unique[:3]  # Max 3 por tema

# Procesar contratos
resultados_contratos = {}

for archivo in sorted(os.listdir(contratos_dir)):
    if not archivo.endswith('.pdf'):
        continue
    
    sindicato = archivo.replace('.pdf', '').replace('_', ' ')
    filepath = os.path.join(contratos_dir, archivo)
    
    pdf = fitz.open(filepath)
    
    # Verificar si tiene texto
    texto_test = pdf[0].get_text()
    es_escaneado = len(texto_test.strip()) < 100
    
    print(f'Procesando: {sindicato}')
    
    if es_escaneado:
        print('  [ESCANEADO - Sin texto]')
        resultados_contratos[sindicato] = {
            'formato_pdf': 'Escaneado',
            'total_paginas': len(pdf),
            'temas': {tema: [{'pagina': 'N/A', 'clausula': 'PDF escaneado - Requiere revisión manual', 'formatos': ['Ver documento']}] for tema in temas.keys()}
        }
    else:
        resultados_contratos[sindicato] = {
            'formato_pdf': 'Texto',
            'total_paginas': len(pdf),
            'temas': {}
        }
        
        for tema, keywords in temas.items():
            clausulas = buscar_clausulas_por_tema(pdf, tema, keywords)
            resultados_contratos[sindicato]['temas'][tema] = clausulas if clausulas else [{'pagina': 'No encontrado', 'clausula': 'No se encontró referencia', 'formatos': ['-']}]
            print(f'  {tema}: {len(clausulas)} referencias')
    
    pdf.close()
    print()

# Crear Excel
print('Creando Excel detallado...')

wb = Workbook()
ws = wb.active
ws.title = 'Clausulas por Sindicato'

# Estilos
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color=WALMART_BLUE, end_color=WALMART_BLUE, fill_type='solid')
subheader_fill = PatternFill(start_color=WALMART_SPARK, end_color=WALMART_SPARK, fill_type='solid')
subheader_font = Font(bold=True, size=10)
green_fill = PatternFill(start_color=GREEN, end_color=GREEN, fill_type='solid')
gray_fill = PatternFill(start_color=GRAY, end_color=GRAY, fill_type='solid')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center')

# Título
ws.merge_cells('A1:G1')
ws['A1'] = 'COMPARATIVO DE CLÁUSULAS - CONTRATOS COLECTIVOS'
ws['A1'].font = Font(bold=True, size=16, color=WALMART_BLUE)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:G2')
ws['A2'] = 'Aniversarios | Excelencia Académica | Uniformes | Casino'
ws['A2'].font = Font(size=12, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Encabezados
headers = ['Sindicato', 'Tema', 'Página', 'Formato/Marca', 'Cláusula / Beneficio', 'Formato PDF', 'Total Págs']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = center_align

# Anchos de columna
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 10
ws.column_dimensions['D'].width = 22
ws.column_dimensions['E'].width = 65
ws.column_dimensions['F'].width = 12
ws.column_dimensions['G'].width = 10

# Datos
row = 5
for sindicato, data in sorted(resultados_contratos.items()):
    first_row_sindicato = row
    
    for tema, clausulas in data['temas'].items():
        for i, clausula_data in enumerate(clausulas):
            # Sindicato (solo primera fila)
            if row == first_row_sindicato:
                ws.cell(row=row, column=1, value=sindicato).font = Font(bold=True)
                ws.cell(row=row, column=6, value=data['formato_pdf'])
                ws.cell(row=row, column=7, value=data['total_paginas'])
            
            # Tema
            cell_tema = ws.cell(row=row, column=2, value=tema if i == 0 else '')
            if i == 0:
                cell_tema.fill = subheader_fill
                cell_tema.font = subheader_font
            
            # Página
            ws.cell(row=row, column=3, value=str(clausula_data['pagina']))
            
            # Formato/Marca
            formatos_str = ', '.join(clausula_data['formatos'])
            ws.cell(row=row, column=4, value=formatos_str)
            
            # Cláusula
            ws.cell(row=row, column=5, value=clausula_data['clausula'])
            
            # Aplicar bordes
            for col in range(1, 8):
                cell = ws.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = wrap_align
            
            row += 1
    
    # Merge celdas de sindicato
    if row - first_row_sindicato > 1:
        ws.merge_cells(f'A{first_row_sindicato}:A{row-1}')
        ws.merge_cells(f'F{first_row_sindicato}:F{row-1}')
        ws.merge_cells(f'G{first_row_sindicato}:G{row-1}')
        ws.cell(row=first_row_sindicato, column=1).alignment = Alignment(vertical='center', horizontal='center')
        ws.cell(row=first_row_sindicato, column=6).alignment = center_align
        ws.cell(row=first_row_sindicato, column=7).alignment = center_align
    
    # Fila separadora
    row += 1

# Ajustar alturas
for r in range(5, row):
    ws.row_dimensions[r].height = 60

# Guardar
wb.save(output_excel)
print(f'Excel guardado: {output_excel}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Clausulas_Contratos_Detallado.xlsx'
shutil.copy2(output_excel, desktop)
print(f'Copiado a: {desktop}')
