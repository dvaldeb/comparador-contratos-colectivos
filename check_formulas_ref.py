# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Cargar archivo de referencia
wb_ref = load_workbook(r'c:\Users\dvaldeb\pupy inteeligente\NC_REF.xlsx')

print('Hojas:', wb_ref.sheetnames)
print()

# Buscar hoja FTEs
ws_ref = None
for sheet in wb_ref.worksheets:
    if 'FTE' in sheet.title:
        ws_ref = sheet
        break

if ws_ref:
    print(f'Hoja encontrada: {ws_ref.title}')
    print(f'Dimensiones: {ws_ref.dimensions}')
    print()
    
    # Ver estructura de encabezados
    print('=== ENCABEZADOS (Filas 1-4) ===')
    for row in range(1, 5):
        vals = []
        for col in range(1, 20):
            val = ws_ref.cell(row=row, column=col).value
            if val:
                vals.append(f'{get_column_letter(col)}:{val}')
        if vals:
            print(f'Fila {row}: {vals}')
    
    print()
    print('=== FORMULAS EN FILA 5 (ejemplo) ===')
    for col in range(1, 45):
        cell = ws_ref.cell(row=5, column=col)
        if cell.value:
            letter = get_column_letter(col)
            # Mostrar si es fórmula o valor
            if isinstance(cell.value, str) and cell.value.startswith('='):
                print(f'{letter}: FORMULA -> {cell.value}')
            else:
                print(f'{letter}: VALOR -> {cell.value}')
    
    print()
    print('=== FORMULAS EN FILAS 10-12 (Max/Min/Mediana) ===')
    for row in range(10, 13):
        label = ws_ref.cell(row=row, column=5).value
        print(f'Fila {row} ({label}):')
        for col in [6, 7, 8, 9, 10, 11]:
            cell = ws_ref.cell(row=row, column=col)
            if cell.value:
                letter = get_column_letter(col)
                print(f'  {letter}: {cell.value}')
else:
    print('No se encontró hoja FTEs')
