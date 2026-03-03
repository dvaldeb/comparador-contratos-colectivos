import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

# Find the exact file
desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'Copia' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

if not path:
    print('ERROR: No se encontr\u00f3 el archivo', flush=True)
    sys.exit(1)

print(f'Archivo: {path}', flush=True)
wb = openpyxl.load_workbook(path)
ws = wb['Value Stream - AT']

# ============================================================
# DATA: 20 etapas (N=4 a N=23) basadas en:
#   - Linea de tiempo negociaciones- Comunicacion operacion1.pptx
#   - Equipos de emergencia (002)1.pptx
#   - Cierre Neg Colectivas 2025 y escenario 2026 1.pptx
#   - Estructura de carpetas de Negociaciones 2025
#   - Archivos de SSMM, N\u00f3minas, Estado de Tiendas
# ============================================================

data = [
    # --- BLOQUE PREPARATORIO ---
    {
        'n': 4,
        'etapa': 'N\u00f3mina y Validaci\u00f3n de Negociadores',
        'proceso': 'Recepci\u00f3n, validaci\u00f3n y cruce de n\u00f3mina de socios negociadores',
        'desc_etapa': (
            'Asegurar que la n\u00f3mina de trabajadores que participar\u00e1n en la negociaci\u00f3n '
            'sea correcta, est\u00e9 vigente y cumpla con los requisitos legales, '
            'habilitando el control de fuero y dotaci\u00f3n por local.'
        ),
        'responsable': 'Relaciones Sindicales, Finanzas (Cami / Juan Millanao), Operaci\u00f3n',
        'desc_proceso': (
            'Se recepciona el listado de negociadores presentado por el sindicato en el proyecto de contrato colectivo. '
            'Se cruza con los registros internos (SAP) para validar vigencia de contrato, cargo y local. '
            'Se identifican trabajadores que pueden incorporarse dentro de 5 d\u00edas. '
            'Se genera la n\u00f3mina de fuero (ej: NOMINA FUERO.28.11.2025.xlsx). '
            'Se comunica a ECO la n\u00f3mina l\u00ednea sin editar v\u00eda SharePoint. '
            'Fuente: Carpeta "N\u00f3minas de Negociadores NOV2025", PPT Slide 1 - Timeline Comunicaci\u00f3n.'
        )
    },
    {
        'n': 5,
        'etapa': 'An\u00e1lisis y Costeo de Propuestas',
        'proceso': 'Evaluaci\u00f3n econ\u00f3mica del proyecto sindical y definici\u00f3n de mandato',
        'desc_etapa': (
            'Cuantificar el impacto econ\u00f3mico de cada cl\u00e1usula propuesta por el sindicato '
            'y preparar los escenarios de respuesta de la empresa, definiendo el mandato '
            'de negociaci\u00f3n con costos claros.'
        ),
        'responsable': 'Finanzas, Compensaciones, Relaciones Sindicales',
        'desc_proceso': (
            'Se costean los beneficios propuestos: reajustes de sueldo base, movilizaci\u00f3n, bonos (BTN, bono operacional, bono correcta marcaci\u00f3n), '
            'asignaciones por zona de pago (cajeros est\u00edmulo, p\u00e9rdida). '
            'Se analiza la eliminaci\u00f3n de cl\u00e1usulas (ej: reajustabilidad SB, bono ROP) y su impacto. '
            'Se eval\u00faan escenarios de costo seg\u00fan adherencia y HC por sindicato. '
            'Fuente: PPT Cierre NC 2025 Slide 4 - Principales modificaciones cla\u00fasulas.'
        )
    },
    # --- BLOQUE NEGOCIACION ---
    {
        'n': 6,
        'etapa': 'Mesa de Negociaci\u00f3n',
        'proceso': 'Desarrollo de sesiones de negociaci\u00f3n con el sindicato',
        'desc_etapa': (
            'Materializar el di\u00e1logo formal entre la empresa y el sindicato para '
            'discutir las cl\u00e1usulas del contrato colectivo, buscando acuerdos que '
            'equilibren intereses de ambas partes dentro del mandato aprobado.'
        ),
        'responsable': 'Relaciones Sindicales, Legal, Finanzas',
        'desc_proceso': (
            'Se realizan sesiones peri\u00f3dicas de negociaci\u00f3n. Se discuten modificaciones clave: '
            'modelo operativo (alcance cl\u00e1usulas jefes, descuento d\u00edas no trabajados en bono responsabilidad), '
            'beneficios (eliminar bono ROP, limitar zona de pago a cajeros, alivianar redacci\u00f3n de beneficios corporativos: '
            '10%, paseo, elecci\u00f3n mejor colaborador, casino, torta cumplea\u00f1os, caja compensaci\u00f3n, t\u00f3mbola), '
            'y costos (acortar brecha entre organizaciones sindicales con bono fijo). '
            'Se levantan actas de cada reuni\u00f3n. '
            'Fuente: PPT Cierre NC 2025 Slides 4-6 - Modificaciones cl\u00e1usulas.'
        )
    },
    {
        'n': 7,
        'etapa': '\u00daltima Oferta',
        'proceso': 'Formulaci\u00f3n y presentaci\u00f3n formal de la \u00faltima oferta al sindicato',
        'desc_etapa': (
            'Formalizar la propuesta final de la empresa, consolidando los acuerdos parciales '
            'y las condiciones econ\u00f3micas definitivas, dentro de los plazos legales del C\u00f3digo del Trabajo.'
        ),
        'responsable': 'Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se consolida la \u00faltima oferta con las cl\u00e1usulas negociadas '
            '(incluye cl\u00e1usulas transitorias como el compromiso de presentar nuevo Bono Operacional Anual '
            'antes de octubre, seg\u00fan PPT Cierre NC Slide 5). '
            'Se redacta formalmente, se valida con legal, y se presenta al sindicato '
            'para someterse a votaci\u00f3n.'
        )
    },
    {
        'n': 8,
        'etapa': 'Votaci\u00f3n',
        'proceso': 'Votaci\u00f3n de la \u00faltima oferta y decisi\u00f3n de huelga',
        'desc_etapa': (
            'Materializar la decisi\u00f3n democr\u00e1tica de los trabajadores '
            'respecto de la \u00faltima oferta, determinando si se acepta el acuerdo o se vota huelga.'
        ),
        'responsable': 'Relaciones Sindicales, Operaci\u00f3n',
        'desc_proceso': (
            'Se coordina el proceso de votaci\u00f3n en los locales afectados. '
            'Se monitorea el qu\u00f3rum legal. Se reciben los resultados oficiales. '
            'Se comunica a la operaci\u00f3n y ECO la fecha probable de huelga si corresponde '
            '(Negociaci\u00f3n Reglada). '
            'Fuente: PPT Timeline Slide 1 - "Fecha Probable de huelga".'
        )
    },
    {
        'n': 9,
        'etapa': 'Buenos Oficios / Mediaci\u00f3n',
        'proceso': 'Mediaci\u00f3n obligatoria por la Inspecci\u00f3n del Trabajo',
        'desc_etapa': (
            'Agotar la instancia legal de mediaci\u00f3n (5 d\u00edas h\u00e1biles) ante la Direcci\u00f3n del Trabajo '
            'antes de hacer efectiva la huelga, buscando un \u00faltimo acuerdo entre las partes.'
        ),
        'responsable': 'Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se participa en las audiencias convocadas por la DT. '
            'Se exploran alternativas de acuerdo. La DT normalmente interviene si '
            'el sindicato se opone a servicios m\u00ednimos. '
            'Se prepara el escenario de huelga en caso de no llegar a acuerdo. '
            'Fuente: PPT Equipos de emergencia Slide 4 - "\u00bfQu\u00e9 nos dice la DT cuando un sindicato se opone?".'
        )
    },
    # --- BLOQUE PREPARACION HUELGA ---
    {
        'n': 10,
        'etapa': 'Servicios M\u00ednimos (SSMM)',
        'proceso': 'Definici\u00f3n y activaci\u00f3n de servicios m\u00ednimos seg\u00fan resoluci\u00f3n vigente',
        'desc_etapa': (
            'Garantizar la continuidad de las funciones estrictamente necesarias '
            'para la protecci\u00f3n de bienes, prevenci\u00f3n sanitaria y seguridad durante la huelga, '
            'conforme a la resoluci\u00f3n de SSMM de Walmart.'
        ),
        'responsable': 'Operaci\u00f3n, Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se aplica la resoluci\u00f3n de SSMM que establece funciones limitadas a: '
            'Guardias de seguridad (vigilancia de instalaciones, NO atenci\u00f3n de p\u00fablico ni revisi\u00f3n de boletas); '
            'Jefe/Encargado de perecibles (retiro de productos descompuestos); '
            'Especialistas carnicero y fiambrero (retiro de alimentos descompuestos/contaminados); '
            'Asistente recepci\u00f3n (traslado para eliminaci\u00f3n de productos). '
            'Se identifican 28 cargos elegibles (desde Carnicero hasta Recepcionista). '
            'Fuente: PPT Timeline Slides 4-5 - Resoluci\u00f3n SSMM y tabla de cargos.'
        )
    },
    {
        'n': 11,
        'etapa': 'Equipos de Emergencia',
        'proceso': 'Conformaci\u00f3n de equipos de emergencia por local y formato',
        'desc_etapa': (
            'Definir qu\u00e9 trabajadores de la n\u00f3mina realizar\u00e1n funciones de servicio m\u00ednimo '
            'en cada local, evaluando utilidad vs riesgo y grado de confianza.'
        ),
        'responsable': 'Operaci\u00f3n (Gerente Tienda), Relaciones Sindicales',
        'desc_proceso': (
            'Se env\u00eda listado de negociadores a cada local. '
            'Los locales se\u00f1alan qui\u00e9nes del proyecto pueden cumplir funciones de SSMM. '
            'Regla por formato: '
            'H\u00edper/Abarrotes Econ\u00f3micos/Mayorista: 2 guardias por turno; '
            'Express/Ekono: 1 guardia por turno. '
            'Todos los formatos: 1 Jefe perecibles + 1 carnicero + 1 fiambrero + 1 asistente recepci\u00f3n. '
            'Se identifica personal finiquitado o sin relaci\u00f3n laboral (requiere finiquito firmado o carta de despido). '
            'Se considera grado de confianza con el trabajador. '
            'Fuente: PPT Timeline Slides 6-7 y PPT Equipos de Emergencia Slides 2-3.'
        )
    },
    # --- BLOQUE HUELGA ---
    {
        'n': 12,
        'etapa': 'Huelga - Activaci\u00f3n',
        'proceso': 'Gesti\u00f3n del inicio y desarrollo de la huelga',
        'desc_etapa': (
            'Administrar el impacto de la huelga en la operaci\u00f3n de tiendas, '
            'resguardando la continuidad del negocio con los equipos de emergencia '
            'y manteniendo control de la situaci\u00f3n en cada local.'
        ),
        'responsable': 'Operaci\u00f3n, Relaciones Sindicales, Gerencia de Mercado',
        'desc_proceso': (
            'Se activa el plan de huelga. Se coordina con los equipos de emergencia '
            'previamente definidos. Se comunica la fecha efectiva de huelga a la operaci\u00f3n. '
            'Se gestionan las tiendas afectadas: en 2025, la NC de FWM afect\u00f3 10.138 HC '
            'y hubo 15 negociaciones colectivas con ~28.000 HC en total. '
            'Fuente: PPT Cierre NC 2025 Slides 2-3 - Timeline y estad\u00edsticas.'
        )
    },
    {
        'n': 13,
        'etapa': 'Estado de Tiendas en Huelga',
        'proceso': 'Monitoreo y reporte diario de tiendas afectadas',
        'desc_etapa': (
            'Mantener visibilidad en tiempo real del estado operacional de cada tienda '
            'durante la huelga, facilitando la toma de decisiones y la reasignaci\u00f3n de recursos.'
        ),
        'responsable': 'Operaci\u00f3n, Relaciones Sindicales, ECO',
        'desc_proceso': (
            'Se consolida informaci\u00f3n diaria por tienda: dotaci\u00f3n presente, adherencia a huelga, \u00e1reas cr\u00edticas. '
            'Se actualizan los par\u00e1metros BI de dotaciones por tribu (L\u00edder, SBA, CM). '
            'Se generan reportes como "Estado de Tiendas FWM NOV2025.xlsx" '
            'y "Operaci\u00f3n Tiendas en Huelga". '
            'Se env\u00edan reportes a gerencia de mercado y operaciones. '
            'Fuente: Carpeta "Estado de Tiendas Huelga (Federaci\u00f3n WM) NOV2025" y archivos de par\u00e1metros BI.'
        )
    },
    {
        'n': 14,
        'etapa': 'Operaci\u00f3n de Tiendas en Huelga',
        'proceso': 'Gesti\u00f3n operacional y reemplazos durante la huelga',
        'desc_etapa': (
            'Asegurar la operaci\u00f3n de las tiendas afectadas mediante redistribuci\u00f3n '
            'de personal, ajuste de par\u00e1metros de dotaci\u00f3n y coordinaci\u00f3n entre locales.'
        ),
        'responsable': 'Operaci\u00f3n, Gerente de Tienda, L\u00edder de Persona',
        'desc_proceso': (
            'Se redistribuyen colaboradores desde tiendas cercanas. '
            'Se ajustan par\u00e1metros BI de dotaciones y negociaciones '
            '(archivos "Parametros BI Dotaciones y Negociaciones 2025" por tribu: L\u00edder, SBA, CM). '
            'Se revisan tiendas de huelga con versi\u00f3n inicial '
            '("Operaci\u00f3n tiendas Negociaci\u00f3n SIL - V. Inicial 07.02.xlsx", '
            '"Todos Huelga - Negociaci\u00f3n 2025.xlsx"). '
            'Fuente: Carpeta "Operaci\u00f3n tiendas Negociaci\u00f3n Mayo 2025".'
        )
    },
    # --- BLOQUE CIERRE ---
    {
        'n': 15,
        'etapa': 'Cierre y Firma del Contrato Colectivo',
        'proceso': 'Redacci\u00f3n final y firma del contrato colectivo',
        'desc_etapa': (
            'Formalizar el acuerdo alcanzado mediante la firma del contrato colectivo, '
            'asegurando que todas las cl\u00e1usulas est\u00e9n debidamente documentadas, '
            'legalizadas y registradas ante la Direcci\u00f3n del Trabajo.'
        ),
        'responsable': 'Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se redacta el texto definitivo del CC incorporando las modificaciones acordadas '
            '(eliminaci\u00f3n reajustabilidad SB, nuevo alcance cl\u00e1usulas jefes, '
            'cl\u00e1usulas transitorias como compromiso de nuevo Bono Operacional). '
            'Se firma por ambas partes. Se registra ante la DT. '
            'Se comunica el cierre a la operaci\u00f3n: RRSS informa, ECO comunica. '
            'Fuente: PPT Timeline Slide 1 - "Firma de CC (T\u00e9rmino de proceso de Neg)".'
        )
    },
    {
        'n': 16,
        'etapa': 'Cambios en Contrato Colectivo',
        'proceso': 'Comunicaci\u00f3n e implementaci\u00f3n de cambios del nuevo CC',
        'desc_etapa': (
            'Asegurar que los cambios del nuevo contrato colectivo sean conocidos '
            'y correctamente aplicados por toda la organizaci\u00f3n, desde operaci\u00f3n hasta compensaciones.'
        ),
        'responsable': 'Relaciones Sindicales (info), ECO (comunicaci\u00f3n), Operaci\u00f3n',
        'desc_proceso': (
            'Se identifican y documentan todos los cambios respecto del CC anterior. '
            'ECO prepara propuesta comunicacional. '
            'Se comunican cambios clave: descuento de d\u00edas no trabajados en bono de responsabilidad y movilizaci\u00f3n, '
            'aclaraci\u00f3n bono operacional, alcance zona de pago (solo cajeros est\u00edmulo y asignaci\u00f3n p\u00e9rdida), '
            'actualizaci\u00f3n de beneficios corporativos (casino actualizado a operador, etc). '
            'Fuente: PPT Timeline Slide 1 - "Cambios de CC", PPT Cierre NC Slide 4.'
        )
    },
    {
        'n': 17,
        'etapa': 'Pago Bono T\u00e9rmino de Negociaci\u00f3n (BTN)',
        'proceso': 'C\u00e1lculo y ejecuci\u00f3n del pago del BTN a colaboradores',
        'desc_etapa': (
            'Materializar el pago del Bono de T\u00e9rmino de Negociaci\u00f3n a los trabajadores '
            'beneficiarios, cumpliendo con los plazos y montos acordados en el contrato colectivo.'
        ),
        'responsable': 'Relaciones Sindicales (info), ECO (comunicaci\u00f3n), Compensaciones, N\u00f3mina',
        'desc_proceso': (
            'Se calcula el BTN seg\u00fan las condiciones del CC firmado. '
            'Se valida la n\u00f3mina de beneficiarios. '
            'Se coordina con n\u00f3mina para la liquidaci\u00f3n y pago. '
            'Se comunica el pago a los colaboradores. '
            'Fuente: PPT Timeline Slide 1 - "Pago de BTN".'
        )
    },
    {
        'n': 18,
        'etapa': 'N\u00f3mina Post-Cierre y Gesti\u00f3n de Fuero',
        'proceso': 'Actualizaci\u00f3n de n\u00f3mina y control de fuero post-firma',
        'desc_etapa': (
            'Actualizar la n\u00f3mina de colaboradores con fuero vigente post-firma del CC '
            'y confirmar la fecha de t\u00e9rmino de fuero (30 d\u00edas despu\u00e9s de la firma), '
            'asegurando cumplimiento legal.'
        ),
        'responsable': 'Finanzas (Cami / Juan Millanao), Relaciones Sindicales, ECO',
        'desc_proceso': (
            'Se genera la n\u00f3mina actualizada de colaboradores con fuero post cierre '
            '("N\u00f3mina l\u00ednea sin editar" en SharePoint). '
            'RRSS confirma la fecha de t\u00e9rmino de fuero (30 d\u00edas post firma del CC). '
            'ECO comunica ambos hitos a la operaci\u00f3n. '
            'Se controlan gestiones de fuero durante y despu\u00e9s de la NC '
            '(ej: NOMINA FUERO 22.12.2025.xlsx). '
            'Fuente: PPT Timeline Slide 1 - "N\u00f3mina actualizada" y "Fecha t\u00e9rmino de Fuero".'
        )
    },
    # --- BLOQUE POST-NC ---
    {
        'n': 19,
        'etapa': 'Implementaci\u00f3n de Beneficios',
        'proceso': 'Parametrizaci\u00f3n y activaci\u00f3n de beneficios del CC en sistemas',
        'desc_etapa': (
            'Materializar los compromisos adquiridos en el contrato colectivo, '
            'asegurando que los beneficios lleguen correctamente a cada trabajador '
            'en los plazos establecidos.'
        ),
        'responsable': 'Compensaciones, N\u00f3mina, Operaci\u00f3n',
        'desc_proceso': (
            'Se parametrizan en sistemas de n\u00f3mina los nuevos beneficios acordados. '
            'Se aplican reajustes salariales, bonos operacionales, BTN, asignaciones. '
            'Se implementan cl\u00e1usulas transitorias (ej: presentar nuevo Bono Operacional antes de octubre). '
            'Se actualizan condiciones corporativas: descuento compras, casino, caja compensaci\u00f3n. '
            'Se verifica la correcta liquidaci\u00f3n en el primer pago. '
            'Fuente: PPT Cierre NC Slides 4-6 - Modificaciones de cl\u00e1usulas.'
        )
    },
    {
        'n': 20,
        'etapa': 'Ecosistema de Escucha Sindical',
        'proceso': 'Activaci\u00f3n y seguimiento del ecosistema de escucha en operaciones',
        'desc_etapa': (
            'Mantener un canal permanente de escucha y relacionamiento sindical '
            'en la operaci\u00f3n, canalizando inquietudes y fortaleciendo la relaci\u00f3n laboral '
            'para prevenir conflictos en el siguiente ciclo de NC.'
        ),
        'responsable': 'Gerente de Mercado, Gerente de Tienda, L\u00edder de Persona (Mercado y Tienda), Prevencionista',
        'desc_proceso': (
            'Se activa el Ecosistema de Escucha Sindical con roles claros: '
            'Gerente Mercado y Gerente Tienda lideran el relacionamiento; '
            'L\u00edder de Persona Mercado y Tienda articulan temas operativos; '
            'Prevencionista es asesor profesional y miembro estable de reuniones sindicales y comit\u00e9 paritario; '
            'Relaciones Sindicales gestiona temas transversales con directiva sindical. '
            'Se genera portafolio de beneficios, insights de voz del colaborador e iniciativas '
            '(Operador PPS, Excelencia Ops, Marcaci\u00f3n, Nuevo Bono Operacional, Estandarizaci\u00f3n Aniversarios, AASS, Casinos). '
            'Fuente: PPT Cierre NC Slides 7-8 - Ecosistema de Escucha.'
        )
    },
    {
        'n': 21,
        'etapa': 'Gesti\u00f3n de Finiquitos NC',
        'proceso': 'Administraci\u00f3n de desvinculaciones asociadas al proceso de NC',
        'desc_etapa': (
            'Gestionar las salidas de colaboradores durante y despu\u00e9s del proceso de NC, '
            'asegurando cumplimiento legal, documental y el correcto tratamiento del fuero.'
        ),
        'responsable': 'Operaci\u00f3n, Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se procesan renuncias voluntarias y desvinculaciones. '
            'Se coordinan finiquitos con resguardo de fuero (finiquito firmado o carta de despido seg\u00fan PPT SSMM). '
            'Se recopila documentaci\u00f3n (contratos, cotizaciones, cartas de despido). '
            'Se gestionan tr\u00e1mites ante la Inspecci\u00f3n del Trabajo. '
            'Se identifica personal finiquitado o sin relaci\u00f3n laboral en pesta\u00f1a aparte del archivo SSMM. '
            'Se actualiza la n\u00f3mina sindical. '
            'Fuente: PPT Timeline Slide 6 - Finiquitados y Gesti\u00f3n Documentos RRSS.'
        )
    },
    {
        'n': 22,
        'etapa': 'Documentaci\u00f3n y Archivo',
        'proceso': 'Gesti\u00f3n documental integral del proceso de NC',
        'desc_etapa': (
            'Preservar toda la documentaci\u00f3n generada durante la negociaci\u00f3n colectiva, '
            'asegurando trazabilidad, organizaci\u00f3n y disponibilidad para futuras referencias '
            'y auditor\u00edas.'
        ),
        'responsable': 'Relaciones Sindicales, ECO, Operaci\u00f3n',
        'desc_proceso': (
            'Se archivan en SharePoint por sindicato y a\u00f1o: '
            'n\u00f3minas de negociadores, SSMM, estados de tiendas, par\u00e1metros BI, '
            'contratos colectivos firmados, finiquitos, actas de reuni\u00f3n, '
            'correspondencia con la DT, y planes comunicacionales. '
            'Se mantiene la estructura de carpetas: Informaci\u00f3n NC / N\u00f3minas / SSMM / Estado de Tiendas / Operaci\u00f3n. '
            'Fuente: Estructura completa de carpetas "FWM Chile Proceso Negociaci\u00f3n colectiva 2025".'
        )
    },
    {
        'n': 23,
        'etapa': 'Seguimiento y Preparaci\u00f3n Pr\u00f3ximo Ciclo',
        'proceso': 'Evaluaci\u00f3n de resultados y preparaci\u00f3n del pr\u00f3ximo ciclo de NC',
        'desc_etapa': (
            'Extraer aprendizajes del proceso de NC finalizado, monitorear '
            'el cumplimiento del CC y preparar la estrategia para el pr\u00f3ximo ciclo '
            'de negociaciones.'
        ),
        'responsable': 'Relaciones Sindicales, Gerencia de Personas, ECO',
        'desc_proceso': (
            'Se realiza an\u00e1lisis retrospectivo: tiempos, costos, adherencia, d\u00edas de huelga, impacto operacional. '
            'Se monitorea el cumplimiento de cl\u00e1usulas transitorias (ej: nuevo Bono Operacional antes de octubre). '
            'Se prepara escenario de pr\u00f3ximo ciclo: '
            'para 2026 se proyectan ~10.000 HC en 10 negociaciones '
            '(Ekono 972 HC, Fenatralid 2.379 HC, Ways Quilicura 749 HC, Fed. Internacional 1.078 HC, '
            'Jefes ex SIL 183 HC, Jefes ex FW 312 HC, Linares SBA 49 HC, Fed. SUNF 1.406 HC, '
            'Ways Paine 87 HC, FSA 2.266 HC). '
            'Se actualizan temas clave: formaci\u00f3n nuevos sindicatos, permisos sindicales, '
            'bono correcta marcaci\u00f3n, derechos adquiridos, actos de injerencia. '
            'Fuente: PPT Cierre NC Slides 9-12 - Timeline 2026 y otros temas RRLL.'
        )
    },
]

# Write data
for row_data in data:
    row = row_data['n'] + 2  # N=0 is row 2, N=4 is row 6, etc.
    ws.cell(row=row, column=1, value=row_data['n'])
    # Column B (VALOR) stays empty
    ws.cell(row=row, column=3, value=row_data['etapa'])
    ws.cell(row=row, column=4, value=row_data['proceso'])
    ws.cell(row=row, column=5, value=row_data['desc_etapa'])
    ws.cell(row=row, column=6, value=row_data['responsable'])
    ws.cell(row=row, column=7, value=row_data['desc_proceso'])

# Save back to the same file
try:
    wb.save(path)
    print(f'\n\u2705 Archivo guardado exitosamente en:', flush=True)
    print(f'   {path}', flush=True)
except PermissionError:
    # If file is open, save to alt path
    alt_path = os.path.join(desktop, 'Copia de Mapeo FTEs  Atenci\u00f3n al Colaborador - Negociacion Colectiva LLENO.xlsx')
    wb.save(alt_path)
    print(f'\n\u26a0\ufe0f Archivo original estaba abierto, se guard\u00f3 como:', flush=True)
    print(f'   {alt_path}', flush=True)

print(f'\nSe llenaron {len(data)} filas (N=4 a N=23)', flush=True)
print('\nResumen de etapas:', flush=True)
for d in data:
    print(f"  N={d['n']:2d} | {d['etapa'][:40]:40s} | {d['responsable'][:50]}", flush=True)
