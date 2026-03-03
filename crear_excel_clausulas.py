# -*- coding: utf-8 -*-
import fitz
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, Fill, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

contratos_dir = r'c:\Users\dvaldeb\pupy inteeligente\contratos'
output_excel = r'c:\Users\dvaldeb\pupy inteeligente\Clausulas_Contratos_Colectivos.xlsx'

# Colores Walmart
WALMART_BLUE = '0053E2'
WALMART_SPARK = 'FFC220'
LIGHT_BLUE = 'E6F0FF'
LIGHT_GRAY = 'F5F5F5'

print('=== CREANDO EXCEL DE CLAUSULAS ===')
print()

# Temas a buscar con palabras clave
temas_config = {
    'Aniversarios': {
        'palabras': ['aniversario', 'años de servicio', 'antiguedad', 'antigüedad', 'quinquenio', 'bono por años'],
        'articulos': []
    },
    'Excelencia Académica': {
        'palabras': ['excelencia acad', 'beca', 'becas', 'estudios', 'escolar', 'educación', 'académico', 'rendimiento escolar'],
        'articulos': []
    },
    'Uniformes': {
        'palabras': ['uniforme', 'uniformes', 'vestimenta', 'vestuario', 'ropa de trabajo', 'indumentaria'],
        'articulos': []
    },
    'Casino': {
        'palabras': ['casino', 'alimentación', 'colación', 'almuerzo', 'comida', 'comedor', 'bono alimentación'],
        'articulos': []
    }
}

def extraer_clausulas(texto, palabras_clave):
    """Extrae párrafos que contengan las palabras clave"""
    clausulas = []
    texto_lower = texto.lower()
    
    # Dividir en párrafos/secciones
    parrafos = re.split(r'\n\s*\n|(?=artículo|(?=art\.\s*\d))', texto, flags=re.IGNORECASE)
    
    for parrafo in parrafos:
        parrafo_lower = parrafo.lower()
        for palabra in palabras_clave:
            if palabra.lower() in parrafo_lower:
                # Limpiar el párrafo
                texto_limpio = ' '.join(parrafo.split())
                if len(texto_limpio) > 30 and texto_limpio not in clausulas:
                    clausulas.append(texto_limpio[:1500])  # Limitar longitud
                break
    
    return clausulas

def buscar_articulo(texto, palabras_clave):
    """Busca el número de artículo relacionado"""
    texto_lower = texto.lower()
    for palabra in palabras_clave:
        # Buscar "Artículo X" cerca de la palabra clave
        patron = r'art[íi]culo\s*(\d+)[^.]*' + re.escape(palabra.lower())
        match = re.search(patron, texto_lower)
        if match:
            return f"Art. {match.group(1)}"
        
        # Buscar en formato "Art. X"
        patron2 = r'art\.\s*(\d+)[^.]*' + re.escape(palabra.lower())
        match2 = re.search(patron2, texto_lower)
        if match2:
            return f"Art. {match2.group(1)}"
    
    return ""

# Procesar cada contrato
resultados = {}

for archivo in sorted(os.listdir(contratos_dir)):
    if not archivo.endswith('.pdf'):
        continue
    
    sindicato = archivo.replace('.pdf', '').replace('_', ' ')
    print(f'Procesando: {sindicato}')
    
    filepath = os.path.join(contratos_dir, archivo)
    pdf = fitz.open(filepath)
    
    # Extraer todo el texto
    texto_completo = ''
    for page in pdf:
        texto_completo += page.get_text() + '\n'
    
    pdf.close()
    
    resultados[sindicato] = {}
    
    for tema, config in temas_config.items():
        clausulas = extraer_clausulas(texto_completo, config['palabras'])
        articulo = buscar_articulo(texto_completo, config['palabras'])
        
        resultados[sindicato][tema] = {
            'articulo': articulo,
            'clausulas': clausulas[:2]  # Max 2 cláusulas por tema
        }
        
        print(f'  {tema}: {len(clausulas)} referencias encontradas')
    
    print()

# Crear Excel
print('Creando Excel...')
wb = Workbook()
ws = wb.active
ws.title = 'Comparativo Cláusulas'

# Estilos
header_font = Font(bold=True, color='FFFFFF', size=12)
header_fill = PatternFill(start_color=WALMART_BLUE, end_color=WALMART_BLUE, fill_type='solid')
subheader_fill = PatternFill(start_color=WALMART_SPARK, end_color=WALMART_SPARK, fill_type='solid')
subheader_font = Font(bold=True, size=11)
alt_fill = PatternFill(start_color=LIGHT_GRAY, end_color=LIGHT_GRAY, fill_type='solid')
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
wrap_alignment = Alignment(wrap_text=True, vertical='top')

# Encabezados
sindicatos = list(resultados.keys())
temas = list(temas_config.keys())

# Fila 1: Título
ws.merge_cells('A1:F1')
ws['A1'] = 'COMPARATIVO DE CLÁUSULAS - CONTRATOS COLECTIVOS'
ws['A1'].font = Font(bold=True, size=16, color=WALMART_BLUE)
ws['A1'].alignment = Alignment(horizontal='center')

# Fila 2: Subtítulo
ws.merge_cells('A2:F2')
ws['A2'] = 'Aniversarios | Excelencia Académica | Uniformes | Casino'
ws['A2'].font = Font(size=12, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Fila 4: Encabezados de columnas
ws['A4'] = 'Sindicato'
ws['B4'] = 'Tema'
ws['C4'] = 'Artículo'
ws['D4'] = 'Cláusula / Beneficio'
ws['E4'] = 'Detalle Adicional'
ws['F4'] = 'Observaciones'

for col in range(1, 7):
    cell = ws.cell(row=4, column=col)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Anchos de columna
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 12
ws.column_dimensions['D'].width = 60
ws.column_dimensions['E'].width = 40
ws.column_dimensions['F'].width = 25

# Llenar datos
row = 5
for sindicato in sindicatos:
    first_row_sindicato = row
    
    for i, tema in enumerate(temas):
        data = resultados[sindicato][tema]
        
        # Sindicato (solo en primera fila de cada sindicato)
        if i == 0:
            ws.cell(row=row, column=1, value=sindicato)
            ws.cell(row=row, column=1).font = Font(bold=True)
        
        # Tema
        ws.cell(row=row, column=2, value=tema)
        ws.cell(row=row, column=2).fill = subheader_fill
        ws.cell(row=row, column=2).font = subheader_font
        
        # Artículo
        ws.cell(row=row, column=3, value=data['articulo'] if data['articulo'] else 'Ver contrato')
        
        # Cláusulas
        if data['clausulas']:
            ws.cell(row=row, column=4, value=data['clausulas'][0][:500])
            if len(data['clausulas']) > 1:
                ws.cell(row=row, column=5, value=data['clausulas'][1][:300])
        else:
            ws.cell(row=row, column=4, value='No se encontró referencia específica')
        
        # Observaciones
        ws.cell(row=row, column=6, value='')
        
        # Aplicar bordes y alineación
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = wrap_alignment
        
        row += 1
    
    # Merge celdas de sindicato
    if row - first_row_sindicato > 1:
        ws.merge_cells(f'A{first_row_sindicato}:A{row-1}')
        ws.cell(row=first_row_sindicato, column=1).alignment = Alignment(vertical='center', horizontal='center')
    
    # Fila separadora
    row += 1

# Ajustar altura de filas
for r in range(5, row):
    ws.row_dimensions[r].height = 80

# Guardar
wb.save(output_excel)
print(f'\nExcel guardado: {output_excel}')

# Copiar al escritorio
import shutil
desktop_path = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Clausulas_Contratos_Colectivos.xlsx'
shutil.copy2(output_excel, desktop_path)
print(f'Copiado a escritorio: {desktop_path}')
