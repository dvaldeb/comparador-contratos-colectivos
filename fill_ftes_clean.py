# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import shutil

# Cerrar Excel si esta abierto
import subprocess
subprocess.run(['taskkill', '/F', '/IM', 'EXCEL.EXE'], capture_output=True)

# Copiar archivo
src = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx'
dst = r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_FINAL2.xlsx'
shutil.copy2(src, dst)

wb = load_workbook(dst)
ws = wb.worksheets[2]  # FTEs

print('=== LIMPIANDO Y LLENANDO HOJA FTEs ===')
print()

# Limpiar filas 5-15 completamente
print('Limpiando filas de datos...')
for row in range(5, 16):
    for col in range(1, 40):
        try:
            ws.cell(row=row, column=col).value = None
        except:
            pass

# Mapeo de etapas a columnas Hr (las columnas FTE tienen formulas)
# F=Et1, H=Et2, J=Et3, L=Et4, N=Et5, P=Et6, R=Et7, T=Et8, V=Et9, X=Et10
etapa_col = {1:6, 2:8, 3:10, 4:12, 5:14, 6:16, 7:18, 8:20, 9:22, 10:24}

# Responsables con sus datos y horas por etapa
responsables = [
    {
        'rut': '11.111.111-1',
        'nombre': 'LPT (Líder Personas Tienda)',
        'cargo': 'Líder Personas Tienda',
        'equipo': 'Operaciones Tienda',
        'jefatura': 'LPM',
        'horas': {1: 2, 2: 3}  # Detección y Registro
    },
    {
        'rut': '22.222.222-2',
        'nombre': 'Gerente de Tienda',
        'cargo': 'Gerente Tienda',
        'equipo': 'Operaciones Tienda',
        'jefatura': 'Squad Lead',
        'horas': {1: 1, 7: 2}  # Detección y Plan Acción
    },
    {
        'rut': '33.333.333-3',
        'nombre': 'Analista RRLL',
        'cargo': 'Analista Relaciones Laborales',
        'equipo': 'RRLL Central',
        'jefatura': 'Líder Comité Multas',
        'horas': {2: 2, 3: 3, 4: 4, 8: 3}  # Registro, Categ, Notif, Seguim
    },
    {
        'rut': '44.444.444-4',
        'nombre': 'ECO (Ejecutivo Clima Org.)',
        'cargo': 'ECO',
        'equipo': 'Gestión Personas',
        'jefatura': 'LPM',
        'horas': {3: 2, 5: 3, 7: 2, 8: 2}  # Categ, Análisis, Plan, Seguim
    },
    {
        'rut': '55.555.555-5',
        'nombre': 'LPM (Líder Personas Mercado)',
        'cargo': 'Líder Personas Mercado',
        'equipo': 'Gestión Personas',
        'jefatura': 'HRBP',
        'horas': {5: 2, 7: 3}  # Análisis y Plan Acción
    },
    {
        'rut': '66.666.666-6',
        'nombre': 'Líder Comité de Multas',
        'cargo': 'Líder Comité Multas',
        'equipo': 'RRLL Central',
        'jefatura': 'Gerente RRLL',
        'horas': {6: 4, 10: 3}  # Sesión Comité y Consolidación
    },
    {
        'rut': '77.777.777-7',
        'nombre': 'Abogado Laboral',
        'cargo': 'Abogado Laboral',
        'equipo': 'Legal',
        'jefatura': 'Gerente Legal',
        'horas': {5: 2, 9: 4}  # Análisis y Resolución
    },
    {
        'rut': '88.888.888-8',
        'nombre': 'Analista Finanzas',
        'cargo': 'Analista Finanzas',
        'equipo': 'Finanzas',
        'jefatura': 'Controller',
        'horas': {9: 2}  # Resolución/Pago
    },
]

print('Escribiendo responsables y horas...')
print()

for i, resp in enumerate(responsables):
    fila = 5 + i
    
    # Datos personales (columnas A-E)
    ws.cell(row=fila, column=1, value=resp['rut'])
    ws.cell(row=fila, column=2, value=resp['nombre'])
    ws.cell(row=fila, column=3, value=resp['cargo'])
    ws.cell(row=fila, column=4, value=resp['equipo'])
    ws.cell(row=fila, column=5, value=resp['jefatura'])
    
    # Horas por etapa
    total_hrs = 0
    horas_str = []
    for etapa_num, hrs in resp['horas'].items():
        col = etapa_col[etapa_num]
        ws.cell(row=fila, column=col, value=hrs)
        total_hrs += hrs
        horas_str.append(f'Et{etapa_num}={hrs}h')
    
    print(f'Fila {fila}: {resp["nombre"]}')
    print(f'         {" | ".join(horas_str)} | Total: {total_hrs} hrs')

print()
print('Guardando...')
wb.save(dst)

# Copiar de vuelta a OneDrive
shutil.copy2(dst, src)
print('Archivo actualizado en OneDrive!')

# Verificar
print()
print('=== VERIFICACION ===')
wb2 = load_workbook(src)
ws2 = wb2.worksheets[2]
for row in range(5, 13):
    nombre = ws2.cell(row=row, column=2).value
    if nombre:
        hrs = [ws2.cell(row=row, column=c).value for c in [6,8,10,12,14,16,18,20,22,24]]
        print(f'Fila {row}: {nombre[:30]}')
        print(f'         Hrs: {hrs}')
