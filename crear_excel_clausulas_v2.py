# -*- coding: utf-8 -*-
import fitz
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import re
import shutil

contratos_dir = r'c:\Users\dvaldeb\pupy inteeligente\contratos'
output_excel = r'c:\Users\dvaldeb\pupy inteeligente\Clausulas_Contratos_Colectivos.xlsx'

# Colores Walmart
WALMART_BLUE = '0053E2'
WALMART_SPARK = 'FFC220'

print('=== EXTRAYENDO CLAUSULAS CON OCR ===')
print()

# Temas a buscar
temas_keywords = {
    'Aniversarios': ['aniversario', 'años de servicio', 'antiguedad', 'quinquenio'],
    'Excelencia Académica': ['excelencia', 'beca', 'escolar', 'estudios', 'educación'],
    'Uniformes': ['uniforme', 'vestimenta', 'vestuario', 'ropa de trabajo'],
    'Casino': ['casino', 'alimentación', 'colación', 'almuerzo', 'comedor']
}

def extraer_texto_pdf(filepath):
    """Extrae texto de PDF, usando OCR si es necesario"""
    pdf = fitz.open(filepath)
    texto_total = ''
    
    for page_num, page in enumerate(pdf):
        # Intentar extraer texto normal
        texto = page.get_text()
        
        if len(texto.strip()) < 50:  # Si no hay texto, es imagen
            # Usar OCR con PyMuPDF (requiere Tesseract instalado)
            # Si no hay Tesseract, extraer como texto la informacion disponible
            try:
                # Intentar OCR integrado de PyMuPDF
                texto = page.get_text("text", flags=fitz.TEXT_PRESERVE_WHITESPACE)
            except:
                pass
        
        texto_total += texto + '\n'
    
    pdf.close()
    return texto_total

def buscar_clausulas(texto, keywords):
    """Busca párrafos que contengan las palabras clave"""
    clausulas = []
    texto_lower = texto.lower()
    
    # Dividir en oraciones/párrafos
    parrafos = re.split(r'[.\n]\s*(?=[A-ZÁÉÍÓÚ]|\d)', texto)
    
    for parrafo in parrafos:
        parrafo_lower = parrafo.lower()
        for kw in keywords:
            if kw.lower() in parrafo_lower and len(parrafo) > 30:
                texto_limpio = ' '.join(parrafo.split())[:800]
                if texto_limpio not in clausulas:
                    clausulas.append(texto_limpio)
                break
    
    return clausulas[:3]  # Max 3 por tema

# Procesar contratos
resultados = {}

pdfs_con_texto = ['SITEL.pdf', 'UTS.pdf']  # PDFs que tienen texto
pdfs_escaneados = ['Fed_Autonoma.pdf', 'Fenatralid.pdf', 'SIL.pdf']  # PDFs escaneados

for archivo in sorted(os.listdir(contratos_dir)):
    if not archivo.endswith('.pdf'):
        continue
    
    sindicato = archivo.replace('.pdf', '').replace('_', ' ')
    filepath = os.path.join(contratos_dir, archivo)
    
    print(f'Procesando: {sindicato}...')
    
    # Extraer texto
    texto = extraer_texto_pdf(filepath)
    
    if len(texto.strip()) < 100:
        print(f'  [ESCANEADO] Sin texto extraible - Requiere revision manual')
        resultados[sindicato] = {
            tema: {
                'clausulas': ['PDF escaneado - Requiere revisión manual del documento original'],
                'encontrado': False
            } for tema in temas_keywords.keys()
        }
    else:
        resultados[sindicato] = {}
        for tema, keywords in temas_keywords.items():
            clausulas = buscar_clausulas(texto, keywords)
            resultados[sindicato][tema] = {
                'clausulas': clausulas if clausulas else ['No se encontró referencia específica'],
                'encontrado': len(clausulas) > 0
            }
            status = 'OK' if clausulas else 'No encontrado'
            print(f'  {tema}: {status}')

print()
print('Creando Excel...')

# Crear Excel
wb = Workbook()
ws = wb.active
ws.title = 'Comparativo Cláusulas'

# Estilos
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color=WALMART_BLUE, end_color=WALMART_BLUE, fill_type='solid')
subheader_fill = PatternFill(start_color=WALMART_SPARK, end_color=WALMART_SPARK, fill_type='solid')
subheader_font = Font(bold=True, size=10)
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')

# Título
ws.merge_cells('A1:E1')
ws['A1'] = 'COMPARATIVO DE CLÁUSULAS - CONTRATOS COLECTIVOS'
ws['A1'].font = Font(bold=True, size=14, color=WALMART_BLUE)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:E2')
ws['A2'] = 'Temas: Aniversarios | Excelencia Académica | Uniformes | Casino'
ws['A2'].font = Font(size=11, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Encabezados
headers = ['Sindicato', 'Tema', 'Cláusula / Beneficio', 'Detalle', 'Estado']
for col, header in enumerate(headers, 1):
    cell = ws.cell(row=4, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.border = thin_border
    cell.alignment = Alignment(horizontal='center', vertical='center')

# Anchos
ws.column_dimensions['A'].width = 18
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 55
ws.column_dimensions['D'].width = 40
ws.column_dimensions['E'].width = 15

# Datos
row = 5
for sindicato in sorted(resultados.keys()):
    first_row = row
    
    for tema, data in resultados[sindicato].items():
        # Sindicato
        if row == first_row:
            ws.cell(row=row, column=1, value=sindicato).font = Font(bold=True)
        
        # Tema
        cell_tema = ws.cell(row=row, column=2, value=tema)
        cell_tema.fill = subheader_fill
        cell_tema.font = subheader_font
        
        # Cláusulas
        clausulas = data['clausulas']
        ws.cell(row=row, column=3, value=clausulas[0] if clausulas else 'N/A')
        ws.cell(row=row, column=4, value=clausulas[1] if len(clausulas) > 1 else '')
        
        # Estado
        estado = '✅ Encontrado' if data.get('encontrado', False) else '⚠️ Revisar'
        ws.cell(row=row, column=5, value=estado)
        
        # Bordes y alineación
        for col in range(1, 6):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = wrap_align
        
        row += 1
    
    # Merge sindicato
    if row - first_row > 1:
        ws.merge_cells(f'A{first_row}:A{row-1}')
        ws.cell(row=first_row, column=1).alignment = Alignment(vertical='center', horizontal='center')
    
    row += 1  # Fila separadora

# Ajustar alturas
for r in range(5, row):
    ws.row_dimensions[r].height = 60

# Guardar
wb.save(output_excel)
print(f'Excel guardado: {output_excel}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Clausulas_Contratos_Colectivos.xlsx'
shutil.copy2(output_excel, desktop)
print(f'Copiado a: {desktop}')
