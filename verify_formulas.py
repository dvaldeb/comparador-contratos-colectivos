# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook(r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx')
ws = wb.worksheets[2]

print('=== VERIFICACION DE FORMULAS ===')
print()

print('FILA 5 (ADP):')
for col in [1, 2, 8, 9, 26, 27]:
    cell = ws.cell(row=5, column=col)
    letra = get_column_letter(col)
    print(f'  {letra}: {cell.value}')

print()
print('FILA 6 (ECO):')
for col in [2, 12, 13, 16, 17, 26, 27]:
    cell = ws.cell(row=6, column=col)
    letra = get_column_letter(col)
    print(f'  {letra}: {cell.value}')

print()
print('FILA 10 (Legal):')
for col in [2, 14, 15, 22, 23, 26, 27]:
    cell = ws.cell(row=10, column=col)
    letra = get_column_letter(col)
    print(f'  {letra}: {cell.value}')

print()
print('Celda de referencia (44 hrs):')
print(f'  AB3: {ws.cell(row=3, column=28).value}')
