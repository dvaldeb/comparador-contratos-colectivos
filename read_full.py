import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

wb = openpyxl.load_workbook(path)
ws = wb['Value Stream - AT']
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=False):
    n = row[0].value
    if n is None:
        continue
    etapa = row[2].value or ''
    proceso = row[3].value or ''
    desc_e = row[4].value or ''
    resp = row[5].value or ''
    desc_p = row[6].value or ''
    print(f'=== N={n} | {etapa} ===', flush=True)
    print(f'  Proceso: {proceso}', flush=True)
    print(f'  Desc Etapa: {desc_e[:200]}', flush=True)
    print(f'  Responsable: {resp}', flush=True)
    print(f'  Desc Proceso: {desc_p[:300]}', flush=True)
    print(flush=True)
