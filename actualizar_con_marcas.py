# -*- coding: utf-8 -*-
import fitz
import os
import re
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
import shutil

print('=== ACTUALIZANDO EXCEL CON MARCAS ===')
print()

# Configuración de temas
temas = {
    'Excelencia Académica': {
        'keywords': ['excelencia acad', 'beca', 'becas', 'escolaridad', 'premio mejor'],
        'fila_excel': 2
    },
    'Aniversarios': {
        'keywords': ['aniversario', 'años de servicio', 'celebrarán su aniversario'],
        'fila_excel': 3
    },
    'Casinos': {
        'keywords': ['casino', 'alimentación', 'colación', 'bono de alimentación'],
        'fila_excel': 4
    },
    'Uniformes': {
        'keywords': ['uniforme', 'ropa de trabajo', 'vestuario', 'vestimenta'],
        'fila_excel': 5
    }
}

# Marcas/Formatos a detectar
marcas_keywords = {
    'Hiper': ['hiper', 'hipermercado', 'híper'],
    'Express': ['express', 'líder express'],
    'SuperBodega (SBA)': ['superbodega', 'sba', 'super bodega', 's.b.a'],
    'Ekono': ['ekono', 'ekóno'],
    'Acuenta': ['acuenta', 'a cuenta'],
    'Central de Distribución': ['centro de distribución', 'c.d.', 'bodega central', 'cd'],
    'Líder': ['líder', 'lider'],
    'Walmart': ['walmart', 'wal-mart', 'wm'],
    'Todas las tiendas': ['todas las tiendas', 'todos los locales', 'todos los establecimientos'],
    'Todos los formatos': ['todos los formatos', 'todas las empresas']
}

# PDFs a procesar
pdfs_config = {
    'SIL_2025.pdf': {
        'path': r'c:\Users\dvaldeb\pupy inteeligente\contratos\SIL_2025.pdf',
        'columna': 4,  # Columna D (SIL)
        'nombre': 'SIL',
        'version': 'CC SIL 2025-2027 Firma 02.07.2025'
    },
    'FED_WM_2025.pdf': {
        'path': r'c:\Users\dvaldeb\pupy inteeligente\contratos\FED_WM_2025.pdf',
        'columna': 3,  # Columna C (FNW)
        'nombre': 'FED WM / FNW',
        'version': 'CC FED WM 15.12.2025 (Versión Definitiva)'
    }
}

def detectar_marcas(texto):
    """Detecta qué marcas se mencionan en el texto"""
    texto_lower = texto.lower()
    marcas_encontradas = []
    
    for marca, keywords in marcas_keywords.items():
        for kw in keywords:
            if kw.lower() in texto_lower:
                if marca not in marcas_encontradas:
                    marcas_encontradas.append(marca)
                break
    
    if not marcas_encontradas:
        return 'General / Todos'
    
    return ', '.join(marcas_encontradas)

def extraer_clausula_completa(pdf, keywords):
    """Busca y extrae la cláusula completa con detección de marcas"""
    for page_num in range(len(pdf)):
        page = pdf[page_num]
        texto = page.get_text()
        texto_lower = texto.lower()
        
        for kw in keywords:
            if kw.lower() in texto_lower:
                idx = texto_lower.find(kw.lower())
                
                # Buscar inicio del artículo
                inicio = max(0, idx - 200)
                for i in range(idx, max(0, idx-200), -1):
                    if re.search(r'\d+[.\-\)]\s*[A-Z]', texto[max(0,i-5):i+5]):
                        inicio = max(0, i-5)
                        break
                
                # Buscar fin del artículo
                fin = min(len(texto), idx + 1500)
                for i in range(idx + 100, min(len(texto), idx + 1500)):
                    if re.search(r'^\s*\d+[.\-\)]\s*[A-Z]', texto[i:i+15]):
                        fin = i
                        break
                
                clausula = texto[inicio:fin].strip()
                clausula = re.sub(r'\n\s*\n+', '\n', clausula)
                
                # Detectar marcas en la cláusula
                marcas = detectar_marcas(clausula)
                
                if len(clausula) > 50:
                    return {
                        'pagina': page_num + 1,
                        'clausula': clausula,
                        'marcas': marcas
                    }
    
    return None

# Cargar Excel original de Downloads
excel_original = r'c:\Users\dvaldeb\Downloads\Cuadro comparativo 4 Beneficios.xlsx'
wb = load_workbook(excel_original)
ws = wb['Hoja1']

print('Estructura actual del Excel:')
print(f'Columnas: A=Beneficio, B=Marca, C=FNW, D=SIL, E=FENATRALID, F=FSA, G=Obs, H=Pag, I=Version')
print()

# Procesar cada PDF
for pdf_name, config in pdfs_config.items():
    print(f'Procesando: {config["nombre"]}')
    pdf = fitz.open(config['path'])
    
    for tema, tema_config in temas.items():
        resultado = extraer_clausula_completa(pdf, tema_config['keywords'])
        
        if resultado:
            fila = tema_config['fila_excel']
            col_clausula = config['columna']
            
            print(f'  {tema}:')
            print(f'    Página: {resultado["pagina"]}')
            print(f'    Marcas: {resultado["marcas"]}')
            
            # Escribir cláusula (columna correspondiente al sindicato)
            ws.cell(row=fila, column=col_clausula, value=resultado['clausula'])
            ws.cell(row=fila, column=col_clausula).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Escribir/actualizar MARCA (columna B = 2)
            marca_actual = ws.cell(row=fila, column=2).value or ''
            nueva_marca = f"{config['nombre']}: {resultado['marcas']}"
            if marca_actual and config['nombre'] not in str(marca_actual):
                ws.cell(row=fila, column=2, value=f"{marca_actual}\n{nueva_marca}")
            else:
                ws.cell(row=fila, column=2, value=nueva_marca if not marca_actual else marca_actual + f"\n{nueva_marca}")
            ws.cell(row=fila, column=2).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Escribir página (columna H = 8)
            pagina_actual = ws.cell(row=fila, column=8).value or ''
            nueva_pagina = f"{config['nombre']}: Pág {resultado['pagina']}"
            if config['nombre'] not in str(pagina_actual):
                if pagina_actual:
                    ws.cell(row=fila, column=8, value=f"{pagina_actual}\n{nueva_pagina}")
                else:
                    ws.cell(row=fila, column=8, value=nueva_pagina)
            ws.cell(row=fila, column=8).alignment = Alignment(wrap_text=True, vertical='top')
            
            # Escribir versión contrato (columna I = 9)
            version_actual = ws.cell(row=fila, column=9).value or ''
            if config['version'] not in str(version_actual):
                if version_actual:
                    ws.cell(row=fila, column=9, value=f"{version_actual}\n{config['version']}")
                else:
                    ws.cell(row=fila, column=9, value=config['version'])
            ws.cell(row=fila, column=9).alignment = Alignment(wrap_text=True, vertical='top')
        else:
            print(f'  {tema}: No encontrado')
    
    pdf.close()
    print()

# Ajustar alturas de fila
for row in range(2, 7):
    ws.row_dimensions[row].height = 180

# Ajustar ancho columna Marca
ws.column_dimensions['B'].width = 35

# Guardar
wb.save(excel_original)
print(f'Excel actualizado: {excel_original}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Cuadro comparativo 4 Beneficios_ACTUALIZADO.xlsx'
shutil.copy2(excel_original, desktop)
print(f'Copiado a: {desktop}')
