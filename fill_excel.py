from openpyxl import load_workbook

# Cargar archivo
wb = load_workbook(r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_temp.xlsx')
ws = wb['Value Stream - AT']

print("Antes de modificar:")
print(f"Fila 3: {[c.value for c in ws[3]]}")

# Etapas del proceso de Comite de Multas
etapas = [
    {
        'n': 1,
        'etapa': 'Deteccion de Contacto Regulatorio',
        'proceso': 'Recepcion de fiscalizacion o requerimiento de la Direccion del Trabajo',
        'descripcion_etapa': 'Identificar y capturar oportunamente los contactos regulatorios de la DT, permitiendo iniciar el proceso de gestion de multas de manera temprana.',
        'responsable': 'LPT / Gerente de Tienda',
        'descripcion_proceso': 'Recibir notificacion de fiscalizacion DT en tienda, registrar datos basicos del inspector, fecha, materia y documentacion solicitada.'
    },
    {
        'n': 2,
        'etapa': 'Registro en Sistema',
        'proceso': 'Documentacion del contacto en sistema de seguimiento',
        'descripcion_etapa': 'Centralizar la informacion del contacto regulatorio en un sistema unico que permita su trazabilidad y seguimiento.',
        'responsable': 'LPT / Analista RRLL',
        'descripcion_proceso': 'Completar formulario de registro con ID unico, fecha, local, mercado, region, materia, tipo de contacto y documentacion adjunta.'
    },
    {
        'n': 3,
        'etapa': 'Categorizacion de la Multa',
        'proceso': 'Clasificacion por materia, gravedad y tipo de infraccion',
        'descripcion_etapa': 'Categorizar cada multa segun su materia para facilitar el analisis de tendencias y definicion de acciones preventivas.',
        'responsable': 'Analista RRLL / ECO',
        'descripcion_proceso': 'Revisar la resolucion o acta de fiscalizacion, identificar materias involucradas, clasificar segun taxonomia definida.'
    },
    {
        'n': 4,
        'etapa': 'Notificacion a Responsables',
        'proceso': 'Comunicacion a LPT, LPM, Gerente y Squad Lead',
        'descripcion_etapa': 'Asegurar que los responsables operacionales esten informados del contacto regulatorio para coordinar respuestas.',
        'responsable': 'Analista RRLL',
        'descripcion_proceso': 'Enviar alerta a responsables con resumen del contacto, materia, documentacion requerida y plazos de respuesta.'
    },
    {
        'n': 5,
        'etapa': 'Analisis y Evaluacion',
        'proceso': 'Revision del caso, documentacion de respaldo y estrategia',
        'descripcion_etapa': 'Evaluar cada caso para determinar la procedencia de la multa, opciones de defensa o subsanacion.',
        'responsable': 'ECO / LPM / Legal',
        'descripcion_proceso': 'Revisar antecedentes, recopilar documentacion de respaldo, evaluar opciones de recurso o subsanacion.'
    },
    {
        'n': 6,
        'etapa': 'Sesion de Comite',
        'proceso': 'Revision en reunion quincenal del Comite de Multas',
        'descripcion_etapa': 'Generar un espacio periodico de revision y decision sobre los casos de multas con visibilidad ejecutiva.',
        'responsable': 'Lider Comite de Multas',
        'descripcion_proceso': 'Preparar presentacion con casos del periodo, revisar multas consolidadas, analizar tendencias, documentar decisiones.'
    },
    {
        'n': 7,
        'etapa': 'Definicion Plan de Accion',
        'proceso': 'Acciones correctivas y preventivas por caso',
        'descripcion_etapa': 'Definir acciones especificas para resolver cada caso y prevenir reincidencias con responsables y plazos claros.',
        'responsable': 'ECO / LPM / Gerente',
        'descripcion_proceso': 'Establecer acciones correctivas, asignar responsables, definir fechas de cumplimiento, documentar en sistema.'
    },
    {
        'n': 8,
        'etapa': 'Seguimiento de Acciones',
        'proceso': 'Monitoreo de cumplimiento del plan de accion',
        'descripcion_etapa': 'Verificar el cumplimiento de las acciones comprometidas, escalando incumplimientos.',
        'responsable': 'Analista RRLL / ECO',
        'descripcion_proceso': 'Monitorear avance de acciones, solicitar evidencia, actualizar estado, escalar casos con incumplimiento.'
    },
    {
        'n': 9,
        'etapa': 'Resolucion y Pago',
        'proceso': 'Cierre de la multa, pago o recurso legal',
        'descripcion_etapa': 'Ejecutar la resolucion final del caso mediante pago de la multa o exito de recurso legal.',
        'responsable': 'Legal / Finanzas',
        'descripcion_proceso': 'Gestionar pago de multa si corresponde, tramitar recursos legales, documentar resolucion final.'
    },
    {
        'n': 10,
        'etapa': 'Consolidacion y Reporte',
        'proceso': 'Documentacion final, reporteria y lecciones aprendidas',
        'descripcion_etapa': 'Generar reportes consolidados de multas por periodo, identificar tendencias y documentar lecciones aprendidas.',
        'responsable': 'Lider Comite de Multas',
        'descripcion_proceso': 'Consolidar datos por formato, mercado y materia, generar reportes ejecutivos, proponer mejoras preventivas.'
    }
]

# Escribir en el Excel
for etapa in etapas:
    row = etapa['n'] + 2  # Fila 3 para N=1, Fila 4 para N=2, etc.
    ws.cell(row=row, column=3, value=etapa['etapa'])
    ws.cell(row=row, column=4, value=etapa['proceso'])
    ws.cell(row=row, column=5, value=etapa['descripcion_etapa'])
    ws.cell(row=row, column=6, value=etapa['responsable'])
    ws.cell(row=row, column=7, value=etapa['descripcion_proceso'])
    print(f"Escrito fila {row}: {etapa['etapa']}")

print("\nDespues de modificar:")
print(f"Fila 3: {[c.value for c in ws[3]]}")

# Guardar
wb.save(r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_temp.xlsx')
print("\nArchivo guardado exitosamente!")
