from openpyxl import load_workbook
import shutil

# Copiar archivo a ubicacion temporal
src = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx'
dst = r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_FINAL.xlsx'

shutil.copy2(src, dst)
print('Archivo copiado')

# Modificar
wb = load_workbook(dst)
ws = wb['Value Stream - AT']

valores = [
    (3, 'Tiempo de captura'),
    (4, 'Completitud registro'),
    (5, 'Precision clasificacion'),
    (6, 'Tiempo notificacion'),
    (7, 'Tasa subsanacion'),
    (8, 'Cobertura casos'),
    (9, 'Acciones definidas'),
    (10, 'Cumplimiento acciones'),
    (11, 'Multas cerradas'),
    (12, 'Tendencia reduccion'),
]

for row, valor in valores:
    ws.cell(row=row, column=2, value=valor)
    print(f'Fila {row}: {valor}')

wb.save(dst)
print('\nArchivo guardado en:', dst)
print('\nVerificacion:')
print('Fila 3:', [c.value for c in ws[3]][:4])

# Copiar de vuelta
shutil.copy2(dst, src)
print('\nCopiado de vuelta a OneDrive!')
