import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
import shutil

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO v2' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

temp = 'temp_check.xlsx'
shutil.copy2(path, temp)
wb = openpyxl.load_workbook(temp)

# === 13 ETAPAS: responsables por etapa ===
print('=== RESPONSABLES POR ETAPA (hoja 13 etapas) ===', flush=True)
ws13 = wb['13 etapas']
for r in range(2, 15):
    n = ws13.cell(row=r, column=1).value
    etapa = ws13.cell(row=r, column=3).value or ''
    resp = ws13.cell(row=r, column=6).value or ''
    print(f'  N={n:2d} | {etapa:45s} | Resp: {resp}', flush=True)

# === FTE's: nombres actuales ===
print('\n=== ROLES ACTUALES EN FTE\'s ===', flush=True)
ws = wb["FTE\u00b4s"]

# Headers de etapas (row 3)
print('\nHeaders columnas (etapas):', flush=True)
stage_cols = [6, 8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30]
from openpyxl.utils import get_column_letter
for c in stage_cols:
    v = ws.cell(row=3, column=c).value or ''
    print(f'  Col {get_column_letter(c)}: {v}', flush=True)

print('\nRoles (filas 5-12):', flush=True)
for r in range(5, 13):
    nombre = ws.cell(row=r, column=2).value or '(vac\u00edo)'
    cargo = ws.cell(row=r, column=3).value or ''
    equipo = ws.cell(row=r, column=4).value or ''
    # Get hours per stage
    hrs = []
    for c in stage_cols:
        h = ws.cell(row=r, column=c).value
        hrs.append(h if h else 0)
    focos = ws.cell(row=r, column=34).value or 0
    print(f'  Row {r}: {nombre:45s} | {equipo:30s} | Hrs: {hrs} | Focos: {focos}', flush=True)

os.remove(temp)
