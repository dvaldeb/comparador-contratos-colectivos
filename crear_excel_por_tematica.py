# -*- coding: utf-8 -*-
import fitz
import os
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import shutil

contratos_dir = r'c:\Users\dvaldeb\pupy inteeligente\contratos'
output_excel = r'c:\Users\dvaldeb\pupy inteeligente\Clausulas_POR_TEMATICA.xlsx'

# Colores Walmart
WALMART_BLUE = '0053E2'
WALMART_SPARK = 'FFC220'
LIGHT_BLUE = 'D6E4FF'
LIGHT_GREEN = 'D4EDDA'
LIGHT_YELLOW = 'FFF3CD'
LIGHT_PURPLE = 'E2D5F1'

print('=== EXTRAYENDO CLAUSULAS POR TEMATICA ===')
print()

# Temas con colores
temas_config = {
    'Aniversarios': {
        'keywords': ['aniversario', 'años de servicio', 'quinquenio', 'años de servicios'],
        'color': 'D6E4FF'  # Azul claro
    },
    'Excelencia Academica': {
        'keywords': ['excelencia acad', 'beca', 'becas escolares', 'premio mejor'],
        'color': 'D4EDDA'  # Verde claro
    },
    'Uniformes': {
        'keywords': ['uniforme', 'ropa de trabajo', 'vestuario', 'vestimenta'],
        'color': 'FFF3CD'  # Amarillo claro
    },
    'Casino': {
        'keywords': ['casino', 'alimentación', 'colación', 'bono alimentaci'],
        'color': 'E2D5F1'  # Morado claro
    }
}

# Formatos/marcas
formatos_keywords = {
    'Hiper': ['hiper', 'hipermercado'],
    'Express': ['express'],
    'SuperBodega': ['superbodega', 'sba', 'bodega'],
    'Ekono': ['ekono'],
    'Central': ['central', 'oficina'],
    'Todos': ['todos los formatos', 'todas las tiendas']
}

def extraer_articulo_completo(texto_pagina, keyword):
    """Extrae el artículo completo que contiene la palabra clave"""
    texto_lower = texto_pagina.lower()
    
    if keyword.lower() not in texto_lower:
        return None
    
    idx = texto_lower.find(keyword.lower())
    
    # Buscar hacia atrás el inicio del artículo
    inicio = max(0, idx - 300)
    for i in range(idx, max(0, idx-300), -1):
        substring = texto_pagina[max(0,i-5):i+5]
        if re.search(r'\d+[.\-\)]\s*[A-ZÁÉÍÓÚ]', substring):
            inicio = max(0, i-5)
            break
    
    # Buscar el final del artículo
    fin = min(len(texto_pagina), idx + 1500)
    for i in range(idx + 100, min(len(texto_pagina), idx + 1500)):
        substring = texto_pagina[i:i+10]
        if re.search(r'^\s*\d+[.\-\)]\s*[A-Z]', substring):
            fin = i
            break
    
    clausula = texto_pagina[inicio:fin].strip()
    clausula = re.sub(r'\n\s*\n+', '\n', clausula)
    
    return clausula

def detectar_formatos(texto):
    """Detecta qué formatos/marcas se mencionan"""
    texto_lower = texto.lower()
    formatos = []
    
    for formato, keywords in formatos_keywords.items():
        for kw in keywords:
            if kw in texto_lower:
                if formato not in formatos:
                    formatos.append(formato)
                break
    
    return ', '.join(formatos) if formatos else 'General'

# Estructura para almacenar resultados POR TEMA
resultados_por_tema = {tema: [] for tema in temas_config.keys()}

# Procesar cada contrato
for archivo in sorted(os.listdir(contratos_dir)):
    if not archivo.endswith('.pdf'):
        continue
    
    sindicato = archivo.replace('.pdf', '').replace('_', ' ')
    filepath = os.path.join(contratos_dir, archivo)
    
    pdf = fitz.open(filepath)
    total_paginas = len(pdf)
    
    # Verificar si tiene texto
    texto_test = pdf[0].get_text()
    es_escaneado = len(texto_test.strip()) < 100
    
    print(f'Procesando: {sindicato} ({total_paginas} págs)')
    
    if es_escaneado:
        print('  [ESCANEADO]')
        for tema in temas_config.keys():
            resultados_por_tema[tema].append({
                'sindicato': sindicato,
                'pagina': 'N/A',
                'formato_marca': 'Ver documento',
                'clausula': 'PDF ESCANEADO - Requiere revisión manual del documento original',
                'tipo_pdf': 'Escaneado',
                'total_pags': total_paginas
            })
    else:
        for tema, config in temas_config.items():
            keywords = config['keywords']
            encontrado = False
            
            for page_num in range(len(pdf)):
                if encontrado:
                    break
                    
                page = pdf[page_num]
                texto = page.get_text()
                
                for kw in keywords:
                    if kw.lower() in texto.lower():
                        clausula = extraer_articulo_completo(texto, kw)
                        
                        if clausula and len(clausula) > 50:
                            formato_marca = detectar_formatos(clausula)
                            
                            resultados_por_tema[tema].append({
                                'sindicato': sindicato,
                                'pagina': page_num + 1,
                                'formato_marca': formato_marca,
                                'clausula': clausula,
                                'tipo_pdf': 'Texto',
                                'total_pags': total_paginas
                            })
                            encontrado = True
                            print(f'  {tema}: Pág {page_num + 1}')
                            break
            
            if not encontrado:
                resultados_por_tema[tema].append({
                    'sindicato': sindicato,
                    'pagina': '-',
                    'formato_marca': '-',
                    'clausula': 'No se encontró referencia específica en el documento',
                    'tipo_pdf': 'Texto',
                    'total_pags': total_paginas
                })
                print(f'  {tema}: No encontrado')
    
    pdf.close()
    print()

# Crear Excel organizado POR TEMA
print('Creando Excel por temática...')

wb = Workbook()
ws = wb.active
ws.title = 'Por Tematica'

# Estilos
header_font = Font(bold=True, color='FFFFFF', size=11)
header_fill = PatternFill(start_color=WALMART_BLUE, end_color=WALMART_BLUE, fill_type='solid')
tema_font = Font(bold=True, size=12, color='FFFFFF')
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap_align = Alignment(wrap_text=True, vertical='top')
center_align = Alignment(horizontal='center', vertical='center')

# Título principal
ws.merge_cells('A1:F1')
ws['A1'] = 'COMPARATIVO DE CLÁUSULAS POR TEMÁTICA'
ws['A1'].font = Font(bold=True, size=18, color=WALMART_BLUE)
ws['A1'].alignment = Alignment(horizontal='center')

ws.merge_cells('A2:F2')
ws['A2'] = 'Contratos Colectivos - Texto extraído directamente de los documentos'
ws['A2'].font = Font(size=11, italic=True)
ws['A2'].alignment = Alignment(horizontal='center')

# Anchos de columna
ws.column_dimensions['A'].width = 16
ws.column_dimensions['B'].width = 8
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 90
ws.column_dimensions['E'].width = 10
ws.column_dimensions['F'].width = 8

row = 4

for tema, config in temas_config.items():
    color_tema = config['color']
    tema_fill = PatternFill(start_color=WALMART_BLUE, end_color=WALMART_BLUE, fill_type='solid')
    row_fill = PatternFill(start_color=color_tema, end_color=color_tema, fill_type='solid')
    
    # Encabezado del tema
    ws.merge_cells(f'A{row}:F{row}')
    cell_tema = ws.cell(row=row, column=1, value=f'📌 {tema.upper()}')
    cell_tema.font = tema_font
    cell_tema.fill = tema_fill
    cell_tema.alignment = Alignment(horizontal='center', vertical='center')
    cell_tema.border = thin_border
    ws.row_dimensions[row].height = 30
    row += 1
    
    # Subencabezados
    subheaders = ['Sindicato', 'Pág.', 'Formato/Marca', 'CLÁUSULA TEXTUAL', 'Tipo', 'Págs']
    for col, header in enumerate(subheaders, 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = Font(bold=True, size=10)
        cell.fill = PatternFill(start_color=WALMART_SPARK, end_color=WALMART_SPARK, fill_type='solid')
        cell.border = thin_border
        cell.alignment = center_align
    row += 1
    
    # Datos del tema
    for item in resultados_por_tema[tema]:
        ws.cell(row=row, column=1, value=item['sindicato']).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(item['pagina']))
        ws.cell(row=row, column=3, value=item['formato_marca'])
        ws.cell(row=row, column=4, value=item['clausula'])
        ws.cell(row=row, column=5, value=item['tipo_pdf'])
        ws.cell(row=row, column=6, value=item['total_pags'])
        
        # Aplicar estilo
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = wrap_align
            cell.fill = row_fill
        
        ws.row_dimensions[row].height = 120
        row += 1
    
    # Espacio entre temas
    row += 1

# Guardar
wb.save(output_excel)
print(f'\nExcel guardado: {output_excel}')

# Copiar al escritorio
desktop = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Clausulas_POR_TEMATICA.xlsx'
shutil.copy2(output_excel, desktop)
print(f'Copiado a: {desktop}')
