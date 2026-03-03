# -*- coding: utf-8 -*-
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import shutil

# Copiar archivo a temporal
src = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx'
dst = r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_FTES.xlsx'
shutil.copy2(src, dst)

wb = load_workbook(dst)
ws_ftes = wb.worksheets[2]  # FTEs

print('=== ACTUALIZANDO HOJA FTEs ===')
print()

# Primero, des-fusionar todas las celdas en el area de encabezados
print('Des-fusionando celdas...')
merged_ranges = list(ws_ftes.merged_cells.ranges)
for merged_range in merged_ranges:
    # Solo des-fusionar las que estan en filas 1-4 y columnas F-AH
    if merged_range.min_row <= 4 and merged_range.min_col >= 6:
        print(f'  Des-fusionando: {merged_range}')
        ws_ftes.unmerge_cells(str(merged_range))

# Limpiar encabezados existentes
print('Limpiando encabezados...')
for row in range(2, 5):
    for col in range(6, 40):
        try:
            ws_ftes.cell(row=row, column=col).value = None
        except:
            pass

# Definir las secciones y etapas del proceso de Comite de Multas
secciones = [
    {
        'nombre': 'DETECCIÓN Y REGISTRO',
        'col_inicio': 6,  # F
        'etapas': [
            'Detección Contacto',  # Etapa 1
            'Registro Sistema',    # Etapa 2
            'Categorización',      # Etapa 3
        ]
    },
    {
        'nombre': 'ANÁLISIS Y NOTIFICACIÓN', 
        'col_inicio': 12,  # L
        'etapas': [
            'Notificación',        # Etapa 4
            'Análisis Caso',       # Etapa 5
        ]
    },
    {
        'nombre': 'GESTIÓN COMITÉ',
        'col_inicio': 16,  # P
        'etapas': [
            'Sesión Comité',       # Etapa 6
            'Plan Acción',         # Etapa 7
            'Seguimiento',         # Etapa 8
        ]
    },
    {
        'nombre': 'CIERRE Y REPORTE',
        'col_inicio': 22,  # V
        'etapas': [
            'Resolución/Pago',     # Etapa 9
            'Consolidación',       # Etapa 10
        ]
    }
]

# Escribir nuevos encabezados
print('Escribiendo nuevos encabezados...')
for seccion in secciones:
    col = seccion['col_inicio']
    print(f"  Seccion '{seccion['nombre']}' en columna {get_column_letter(col)}")
    
    # Fila 2: Nombre de la seccion
    ws_ftes.cell(row=2, column=col, value=seccion['nombre'])
    
    # Fila 3: Nombres de etapas (una cada 2 columnas: Hr, FTE)
    for i, etapa in enumerate(seccion['etapas']):
        etapa_col = col + (i * 2)
        ws_ftes.cell(row=3, column=etapa_col, value=etapa)
        print(f"    Etapa '{etapa}' en {get_column_letter(etapa_col)}")
        
        # Fila 4: Hr y FTE
        ws_ftes.cell(row=4, column=etapa_col, value='Hr')
        ws_ftes.cell(row=4, column=etapa_col + 1, value='FTE')

# Totales
ws_ftes.cell(row=3, column=26, value='Focos/Proyectos')  # Z
ws_ftes.cell(row=4, column=26, value='Hr')
ws_ftes.cell(row=4, column=27, value='FTE')

ws_ftes.cell(row=3, column=28, value='TOTAL HRS SEM')  # AB
ws_ftes.cell(row=3, column=29, value='TOTAL FTE')      # AC

print()
print('Guardando archivo...')
wb.save(dst)

# Copiar de vuelta
shutil.copy2(dst, src)
print('Archivo actualizado y copiado a OneDrive!')

print()
print('=== VERIFICACION ===')
wb2 = load_workbook(src)
ws2 = wb2.worksheets[2]
print('Fila 2 (Secciones):')
for col in range(6, 30):
    val = ws2.cell(row=2, column=col).value
    if val:
        print(f'  {get_column_letter(col)}: {val}')
print('Fila 3 (Etapas):')
for col in range(6, 30):
    val = ws2.cell(row=3, column=col).value
    if val:
        print(f'  {get_column_letter(col)}: {val}')
print('Fila 4 (Hr/FTE):')
for col in range(6, 30):
    val = ws2.cell(row=4, column=col).value
    if val:
        print(f'  {get_column_letter(col)}: {val}')
