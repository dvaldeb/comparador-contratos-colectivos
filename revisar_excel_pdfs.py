# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import fitz
import os

# 1. Revisar estructura del Excel
print('=== ESTRUCTURA DEL EXCEL ===')
wb = load_workbook(r'c:\Users\dvaldeb\pupy inteeligente\Cuadro_Comparativo.xlsx')
print(f'Hojas: {wb.sheetnames}')
print()

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'Hoja: {sheet_name}')
    print(f'Dimensiones: {ws.dimensions}')
    
    # Mostrar primeras filas
    print('Contenido:')
    for row in range(1, min(10, ws.max_row + 1)):
        valores = []
        for col in range(1, min(10, ws.max_column + 1)):
            val = ws.cell(row=row, column=col).value
            if val:
                valores.append(f'{val}'[:30])
        if valores:
            print(f'  Fila {row}: {valores}')
    print()

wb.close()

# 2. Verificar PDFs
print('=== VERIFICANDO PDFs ===')
pdfs = [
    r'c:\Users\dvaldeb\pupy inteeligente\contratos\SIL_2025.pdf',
    r'c:\Users\dvaldeb\pupy inteeligente\contratos\FED_WM_2025.pdf'
]

for pdf_path in pdfs:
    nombre = os.path.basename(pdf_path)
    pdf = fitz.open(pdf_path)
    
    # Extraer texto de primera página
    texto = pdf[0].get_text()
    tiene_texto = len(texto.strip()) > 100
    
    print(f'{nombre}:')
    print(f'  Páginas: {len(pdf)}')
    print(f'  Tiene texto: {"Sí" if tiene_texto else "No (escaneado)"}')
    print(f'  Caracteres pág 1: {len(texto)}')
    
    pdf.close()
    print()
