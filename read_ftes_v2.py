import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO v2' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

print(f'Archivo: {path}', flush=True)

import shutil
temp = 'temp_v2.xlsx'
shutil.copy2(path, temp)
wb = openpyxl.load_workbook(temp)

for sn in wb.sheetnames:
    if 'FTE' in sn.upper():
        ws = wb[sn]
        print(f'\n=== Sheet: "{sn}" ===', flush=True)
        print(f'Max row: {ws.max_row}, Max col: {ws.max_column}', flush=True)
        
        # Print ALL cells with values for rows 1-16
        for r in range(1, ws.max_row + 1):
            row_data = []
            for c in range(1, ws.max_column + 1):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    col_letter = openpyxl.utils.get_column_letter(c)
                    row_data.append(f'{col_letter}{r}={repr(val)[:60]}')
            if row_data:
                print(f'  Row {r}: {row_data}', flush=True)
        
        # Print merged cells
        print(f'\n  Merged cells: {list(ws.merged_cells.ranges)}', flush=True)
