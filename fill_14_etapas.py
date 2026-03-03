import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl

desktop = r'C:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio'
path = None
for f in os.listdir(desktop):
    if 'COMPLETO' in f and 'Mapeo' in f:
        path = os.path.join(desktop, f)
        break

print(f'Archivo: {path}', flush=True)
wb = openpyxl.load_workbook(path)
ws = wb['Value Stream - AT']

# ============================================================
# 14 ETAPAS CONSOLIDADAS (N=0 a N=13)
# Fusiones:
#   N=1 nuevo = viejo N=1 + N=3 (Inicio Formal + Fuero)
#   N=4 nuevo = viejo N=6 + N=16 (Análisis Financiero + Costeo Final)
#   N=6 nuevo = viejo N=7 + N=8 (\u00daltima Oferta + Votaci\u00f3n)
#   N=9 nuevo = viejo N=11 + N=12 + N=13 (Huelga + Reemplazos + Estado Tiendas)
#   N=10 nuevo = viejo N=14 + N=15 + N=19 (Bajas + Cierre + Registro)
#   N=11 nuevo = viejo N=17 + N=18 (Implementaci\u00f3n + Reincorporaci\u00f3n)
#   N=12 nuevo = viejo N=20 + N=21 (Comunicaci\u00f3n Post + Seguimiento)
# ============================================================

data = [
    {
        'n': 0,
        'etapa': 'Planificaci\u00f3n',
        'proceso': 'Preparaci\u00f3n estrat\u00e9gica de la negociaci\u00f3n',
        'desc_etapa': (
            'Permitir a la organizaci\u00f3n anticiparse al proceso de negociaci\u00f3n colectiva '
            'mediante la definici\u00f3n de estrategia, an\u00e1lisis de impactos y levantamiento '
            'de informaci\u00f3n clave, asegurando preparaci\u00f3n legal, financiera y operacional.'
        ),
        'responsable': 'Relaciones Sindicales, Finanzas',
        'desc_proceso': (
            'Se identifican sindicatos y fechas de vencimiento de contratos colectivos, '
            'se levanta la dotaci\u00f3n sindicalizada, se analizan convenios vigentes, '
            'se realiza costeo preliminar de beneficios y se define la estrategia negociadora '
            'junto al equipo legal y de compensaciones.'
        )
    },
    {
        'n': 1,
        'etapa': 'Inicio Formal y Fuero',
        'proceso': 'Recepci\u00f3n de proyecto, activaci\u00f3n de negociaci\u00f3n y gesti\u00f3n de fuero',
        'desc_etapa': (
            'Formalizar el inicio del proceso de negociaci\u00f3n colectiva, habilitando su '
            'revisi\u00f3n legal y an\u00e1lisis econ\u00f3mico, y resguardar los derechos de fuero '
            'de los trabajadores involucrados (10 d\u00edas antes de la presentaci\u00f3n del proyecto '
            'hasta 30 d\u00edas despu\u00e9s de la firma del CC), asegurando cumplimiento normativo '
            'y mitigando riesgos legales.'
        ),
        'responsable': 'Relaciones Sindicales, Finanzas, Operaci\u00f3n',
        'desc_proceso': (
            'Se recepciona el proyecto de contrato colectivo (negociaci\u00f3n reglada) o se formaliza '
            'la solicitud de negociaci\u00f3n (no reglada). Se valida jur\u00eddicamente el documento '
            'y se analizan sus cl\u00e1usulas econ\u00f3micas y normativas. '
            'Simult\u00e1neamente se activa el control de fuero: se identifica a los trabajadores '
            'con fuero vigente, se comunica a la operaci\u00f3n las restricciones de desvinculaci\u00f3n '
            'y se monitorea el cumplimiento durante todo el proceso.'
        )
    },
    {
        'n': 2,
        'etapa': 'N\u00f3mina y Validaci\u00f3n',
        'proceso': 'Validaci\u00f3n de n\u00f3mina negociadora y socios',
        'desc_etapa': (
            'Asegurar que la n\u00f3mina de trabajadores involucrados en la negociaci\u00f3n sea correcta, '
            'actualizada y cumpla con los requisitos legales para participar del proceso, '
            'habilitando el control de dotaci\u00f3n por local.'
        ),
        'responsable': 'Relaciones Sindicales, Operaci\u00f3n',
        'desc_proceso': (
            'Se recepciona la n\u00f3mina de negociadores del sindicato, se valida contra los registros '
            'de SAP/sistemas internos, se identifican colaboradores con fuero, se cruza con '
            'dotaci\u00f3n por local y se comunica a los l\u00edderes de persona en tienda para control. '
            'Se identifican trabajadores que pueden incorporarse dentro de 5 d\u00edas h\u00e1biles '
            'y se genera la n\u00f3mina de fuero consolidada.'
        )
    },
    {
        'n': 3,
        'etapa': 'Comunicaci\u00f3n del Proceso',
        'proceso': 'Gesti\u00f3n comunicacional de la negociaci\u00f3n colectiva',
        'desc_etapa': (
            'Alinear a los distintos niveles de la organizaci\u00f3n respecto del inicio y desarrollo '
            'de la negociaci\u00f3n, entregando lineamientos claros de actuaci\u00f3n y resguardos legales.'
        ),
        'responsable': 'Relaciones Sindicales (info), ECO (comunicaci\u00f3n), Operaciones, Comunicaciones Internas',
        'desc_proceso': (
            'Se comunica el inicio de la negociaci\u00f3n a la operaci\u00f3n, se presentan los alcances '
            'del proyecto sindical, se informan restricciones legales y se entregan directrices '
            'para la correcta gesti\u00f3n del proceso en locales y \u00e1reas de apoyo. '
            'RRSS es responsable de la informaci\u00f3n y ECO de comunicar. '
            'Se comunican hitos clave: inicio de fuero, fecha probable de huelga, '
            'firma del CC, cambios del CC, pago del BTN y t\u00e9rmino de fuero.'
        )
    },
    {
        'n': 4,
        'etapa': 'An\u00e1lisis Financiero y Costeo',
        'proceso': 'Costeo, evaluaci\u00f3n econ\u00f3mica de propuestas y ajuste presupuestario',
        'desc_etapa': (
            'Cuantificar el impacto econ\u00f3mico de cada propuesta sindical y de la empresa, '
            'tanto durante la negociaci\u00f3n como tras la firma del CC, permitiendo '
            'decisiones informadas y el correcto ajuste presupuestario.'
        ),
        'responsable': 'Finanzas, Compensaciones, Relaciones Sindicales',
        'desc_proceso': (
            'Se costean los beneficios propuestos (reajustes, bonos, SSMM), se analizan escenarios '
            'de costo por dotaci\u00f3n afectada, se proyectan impactos presupuestarios anuales '
            'y se preparan reportes comparativos respecto del contrato vigente. '
            'Post-firma: se costea el CC firmado por beneficio y por local, se compara con el '
            'presupuesto asignado, se generan provisiones contables y se comunica el impacto '
            'financiero a la gerencia.'
        )
    },
    {
        'n': 5,
        'etapa': 'Mesa de Negociaci\u00f3n',
        'proceso': 'Desarrollo de sesiones de negociaci\u00f3n con el sindicato',
        'desc_etapa': (
            'Generar el espacio formal de di\u00e1logo entre la empresa y el sindicato para discutir '
            'las cl\u00e1usulas del contrato colectivo, buscando acuerdos que equilibren '
            'los intereses de ambas partes.'
        ),
        'responsable': 'Relaciones Sindicales, Legal, Finanzas',
        'desc_proceso': (
            'Se realizan sesiones peri\u00f3dicas de negociaci\u00f3n, se analizan propuestas y '
            'contrapropuestas sindicales, se levantan actas de cada reuni\u00f3n, se eval\u00faan '
            'impactos econ\u00f3micos de las solicitudes y se coordinan respuestas con las \u00e1reas '
            'de soporte (finanzas, legal, compensaciones).'
        )
    },
    {
        'n': 6,
        'etapa': '\u00daltima Oferta y Votaci\u00f3n',
        'proceso': 'Formulaci\u00f3n de \u00faltima oferta, presentaci\u00f3n y votaci\u00f3n',
        'desc_etapa': (
            'Formalizar la propuesta final de la empresa al sindicato y materializar la decisi\u00f3n '
            'democr\u00e1tica de los trabajadores, determinando si se acepta el acuerdo o se vota huelga.'
        ),
        'responsable': 'Relaciones Sindicales, Legal, Operaci\u00f3n',
        'desc_proceso': (
            'Se consolida la \u00faltima oferta considerando los acuerdos parciales alcanzados, '
            'se redacta formalmente el documento y se presenta al sindicato en los plazos legales. '
            'Se coordina el proceso de votaci\u00f3n en los locales afectados, se monitorea el qu\u00f3rum '
            'legal, se reciben los resultados oficiales y se comunican internamente para activar '
            'el plan correspondiente (firma o huelga). Se informa a ECO la fecha probable de huelga '
            'si corresponde (Negociaci\u00f3n Reglada).'
        )
    },
    {
        'n': 7,
        'etapa': 'Buenos Oficios / Mediaci\u00f3n',
        'proceso': 'Mediaci\u00f3n obligatoria por la Inspecci\u00f3n del Trabajo',
        'desc_etapa': (
            'Facilitar un \u00faltimo espacio de mediaci\u00f3n obligatoria a trav\u00e9s de la Inspecci\u00f3n del Trabajo '
            'antes del inicio de la huelga, buscando un acuerdo entre las partes.'
        ),
        'responsable': 'Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se participa en las audiencias de buenos oficios convocadas por la Direcci\u00f3n del Trabajo '
            '(5 d\u00edas h\u00e1biles), se exploran alternativas de acuerdo, se eval\u00faan nuevas propuestas '
            'y se prepara el escenario de huelga en caso de no llegar a acuerdo.'
        )
    },
    {
        'n': 8,
        'etapa': 'SSMM y Equipos de Emergencia',
        'proceso': 'Definici\u00f3n de servicios m\u00ednimos y conformaci\u00f3n de equipos de emergencia',
        'desc_etapa': (
            'Garantizar la continuidad de las operaciones esenciales durante la huelga, '
            'conforme a la resoluci\u00f3n de SSMM vigente, definiendo qu\u00e9 trabajadores '
            'cumplir\u00e1n funciones de servicio m\u00ednimo en cada local por formato.'
        ),
        'responsable': 'Operaci\u00f3n (Gerente Tienda), Relaciones Sindicales, Legal',
        'desc_proceso': (
            'Se aplica la resoluci\u00f3n de SSMM de Walmart: funciones limitadas a guardias de seguridad '
            '(H\u00edper/Mayorista: 2 por turno; Express/Ekono: 1 por turno), '
            'Jefe/Encargado de perecibles, especialistas carnicero y fiambrero, '
            'y asistente de recepci\u00f3n. Funci\u00f3n exclusiva: retiro de productos descompuestos. '
            'Se env\u00eda n\u00f3mina a locales para que se\u00f1alen qui\u00e9nes del proyecto pueden cumplir SSMM '
            '(evaluando utilidad vs riesgo y grado de confianza). '
            'Se identifican finiquitados/sin relaci\u00f3n laboral en pesta\u00f1a aparte. '
            'Se notifica al sindicato y a la DT.'
        )
    },
    {
        'n': 9,
        'etapa': 'Huelga y Gesti\u00f3n Operacional',
        'proceso': 'Gesti\u00f3n de la huelga, reemplazos, dotaci\u00f3n y monitoreo de tiendas',
        'desc_etapa': (
            'Administrar integralmente el impacto de la huelga: monitorear el estado de cada tienda, '
            'gestionar reemplazos y redistribuci\u00f3n de personal, y mantener visibilidad '
            'en tiempo real para la toma de decisiones.'
        ),
        'responsable': 'Operaci\u00f3n, Relaciones Sindicales, Gerencia de Mercado, Gerencia de Tienda',
        'desc_proceso': (
            'Se monitorea el estado de cada tienda afectada: dotaci\u00f3n presente, adherencia, \u00e1reas cr\u00edticas. '
            'Se gestionan par\u00e1metros de dotaci\u00f3n en BI por tribu (L\u00edder, SBA, CM). '
            'Se coordinan equipos de apoyo entre locales y se redistribuyen colaboradores '
            'desde tiendas cercanas. Se activan procesos de contrataci\u00f3n temporal si corresponde. '
            'Se controla el ingreso/salida de trabajadores en huelga. '
            'Se reporta diariamente el estado operacional a gerencia de mercado y operaciones '
            '(reportes tipo "Estado de Tiendas FWM", "Operaci\u00f3n Tiendas en Huelga").'
        )
    },
    {
        'n': 10,
        'etapa': 'Cierre, Firma y Registro Legal',
        'proceso': 'Cierre de negociaci\u00f3n, firma del CC, gesti\u00f3n de bajas y registro ante la DT',
        'desc_etapa': (
            'Formalizar el acuerdo mediante la firma del contrato colectivo, gestionar '
            'las desvinculaciones asociadas al proceso, y cumplir con la obligaci\u00f3n legal '
            'de registrar el CC ante la Direcci\u00f3n del Trabajo.'
        ),
        'responsable': 'Relaciones Sindicales, Legal, Operaci\u00f3n',
        'desc_proceso': (
            'Se redacta el texto definitivo del CC, se revisa jur\u00eddicamente cada cl\u00e1usula, '
            'se firma por ambas partes y se distribuyen copias a las \u00e1reas involucradas. '
            'Se procesan renuncias voluntarias y finiquitos con resguardo de fuero, '
            'se recopila documentaci\u00f3n de soporte y se gestionan tr\u00e1mites ante la Inspecci\u00f3n del Trabajo. '
            'Se prepara y env\u00eda el CC firmado para dep\u00f3sito en la DT dentro del plazo legal '
            '(5 d\u00edas h\u00e1biles), se obtiene constancia de registro y se archiva el comprobante.'
        )
    },
    {
        'n': 11,
        'etapa': 'Implementaci\u00f3n, BTN y Reincorporaci\u00f3n',
        'proceso': 'Activaci\u00f3n de beneficios, pago BTN y reincorporaci\u00f3n post-huelga',
        'desc_etapa': (
            'Materializar los compromisos del CC (beneficios, bonos, reajustes), '
            'ejecutar el pago del Bono de T\u00e9rmino de Negociaci\u00f3n y facilitar '
            'el retorno ordenado de los trabajadores post-huelga.'
        ),
        'responsable': 'Compensaciones, N\u00f3mina, Operaci\u00f3n, Gerencia de Tienda, L\u00edder de Persona',
        'desc_proceso': (
            'Se parametrizan los nuevos beneficios en sistemas de n\u00f3mina, se aplican reajustes '
            'salariales, se activan bonos y asignaciones, se verifican liquidaciones. '
            'Se calcula y ejecuta el pago del BTN seg\u00fan condiciones del CC. '
            'Se coordina el reingreso de trabajadores a funciones habituales, se actualizan '
            'planificaciones y turnos, se resuelven situaciones pendientes '
            'y se monitorea el clima laboral post-huelga. '
            'Se actualiza la n\u00f3mina de fuero post-cierre (30 d\u00edas despu\u00e9s de la firma) '
            'y se confirma fecha de t\u00e9rmino de fuero.'
        )
    },
    {
        'n': 12,
        'etapa': 'Comunicaci\u00f3n Post-NC y Seguimiento',
        'proceso': 'Comunicaci\u00f3n de resultados y monitoreo de cumplimiento del CC',
        'desc_etapa': (
            'Informar a toda la organizaci\u00f3n sobre los resultados y compromisos adquiridos, '
            'y verificar que las cl\u00e1usulas del CC se cumplan sostenidamente, '
            'previniendo conflictos y fortaleciendo la relaci\u00f3n laboral.'
        ),
        'responsable': 'Comunicaciones Internas, Relaciones Sindicales, Operaci\u00f3n, Compensaciones',
        'desc_proceso': (
            'Se elaboran comunicados de cierre, se realizan reuniones de bajada con gerentes '
            'de mercado y l\u00edderes de persona, se comunican los nuevos beneficios '
            'y se actualizan manuales de referencia. '
            'Se realiza seguimiento peri\u00f3dico al cumplimiento de cl\u00e1usulas, '
            'se resuelven consultas de tiendas sobre aplicaci\u00f3n de beneficios, '
            'se monitorean indicadores de clima laboral y se identifican alertas tempranas '
            'de incumplimiento para gesti\u00f3n preventiva.'
        )
    },
    {
        'n': 13,
        'etapa': 'Evaluaci\u00f3n, Documentaci\u00f3n y Aprendizaje',
        'proceso': 'Evaluaci\u00f3n del proceso, gesti\u00f3n documental y lecciones aprendidas',
        'desc_etapa': (
            'Extraer aprendizajes del proceso de NC, preservar toda la documentaci\u00f3n generada '
            'y preparar la estrategia para el pr\u00f3ximo ciclo de negociaciones.'
        ),
        'responsable': 'Relaciones Sindicales, Gerencia de Personas, ECO',
        'desc_proceso': (
            'Se realiza an\u00e1lisis retrospectivo del proceso (tiempos, costos, resultados, '
            'd\u00edas de huelga, impacto operacional). Se documentan lecciones aprendidas '
            'y se identifican mejoras para el pr\u00f3ximo ciclo. '
            'Se archivan en SharePoint por sindicato y a\u00f1o: n\u00f3minas, SSMM, estados de tiendas, '
            'par\u00e1metros BI, CC firmados, finiquitos, actas y correspondencia con la DT. '
            'Se actualiza la estrategia de relacionamiento sindical '
            'y se prepara el escenario del pr\u00f3ximo ciclo de NC.'
        )
    },
]

# --- WRITE DATA ---
# First, clear ALL rows from 2 to max_row
for row in range(2, ws.max_row + 1):
    for col in range(1, 8):  # Columns A to G
        ws.cell(row=row, column=col, value=None)

# Write the 14 stages
for row_data in data:
    row = row_data['n'] + 2  # N=0 -> row 2
    ws.cell(row=row, column=1, value=row_data['n'])
    # Column B (VALOR) stays empty
    ws.cell(row=row, column=3, value=row_data['etapa'])
    ws.cell(row=row, column=4, value=row_data['proceso'])
    ws.cell(row=row, column=5, value=row_data['desc_etapa'])
    ws.cell(row=row, column=6, value=row_data['responsable'])
    ws.cell(row=row, column=7, value=row_data['desc_proceso'])

# Save
try:
    wb.save(path)
    print(f'\n\u2705 Guardado exitosamente en:\n   {path}', flush=True)
except PermissionError:
    alt = path.replace('COMPLETO', 'COMPLETO 14 ETAPAS')
    wb.save(alt)
    print(f'\n\u26a0\ufe0f Archivo original abierto, guardado como:\n   {alt}', flush=True)

print(f'\n\u2705 Se escribieron {len(data)} etapas (N=0 a N=13)', flush=True)
print(f'\u2705 Se limpiaron las filas sobrantes (N=14 a N=22)', flush=True)
print('\nResumen:', flush=True)
for d in data:
    print(f"  N={d['n']:2d} | {d['etapa']}", flush=True)
