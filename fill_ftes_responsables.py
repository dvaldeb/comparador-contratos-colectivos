# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import shutil

# Copiar archivo
src = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx'
dst = r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_FTES2.xlsx'
shutil.copy2(src, dst)

wb = load_workbook(dst)
ws_vs = wb['Value Stream - AT']
ws_ftes = wb.worksheets[2]  # FTEs

print('=== RESPONSABLES DEL VALUE STREAM ===')
print()

# Leer responsables del Value Stream
etapas_responsables = []
for row in range(3, 13):
    n = ws_vs.cell(row=row, column=1).value
    etapa = ws_vs.cell(row=row, column=3).value
    responsable = ws_vs.cell(row=row, column=6).value
    if n and etapa and responsable:
        etapas_responsables.append({
            'num': n,
            'etapa': etapa,
            'responsable': responsable
        })
        print(f'Etapa {n}: {responsable}')

print()

# Extraer roles unicos
roles_unicos = set()
for er in etapas_responsables:
    # Separar roles compuestos (ej: "LPT / Gerente de Tienda")
    for rol in er['responsable'].split(' / '):
        roles_unicos.add(rol.strip())

print('=== ROLES UNICOS ===')
for rol in sorted(roles_unicos):
    print(f'  - {rol}')

print()

# Mapeo de etapas a columnas en FTEs
# Basado en la estructura que creamos:
# F=Deteccion, H=Registro, J=Categorizacion, L=Notificacion, N=Analisis
# P=Sesion Comite, R=Plan Accion, T=Seguimiento, V=Resolucion, X=Consolidacion
etapa_a_columna = {
    1: 6,   # F - Deteccion
    2: 8,   # H - Registro
    3: 10,  # J - Categorizacion
    4: 12,  # L - Notificacion
    5: 14,  # N - Analisis
    6: 16,  # P - Sesion Comite
    7: 18,  # R - Plan Accion
    8: 20,  # T - Seguimiento
    9: 22,  # V - Resolucion
    10: 24, # X - Consolidacion
}

# Definir responsables principales (simplificados) y sus participaciones
# Cada rol tiene las etapas donde participa y horas estimadas por etapa
responsables_data = [
    {
        'nombre': 'LPT (Lider Personas Tienda)',
        'cargo': 'Lider Personas Tienda',
        'equipo': 'Operaciones Tienda',
        'jefatura': 'LPM',
        'horas': {1: 2, 2: 3}  # Etapas 1 y 2
    },
    {
        'nombre': 'Gerente de Tienda',
        'cargo': 'Gerente Tienda',
        'equipo': 'Operaciones Tienda',
        'jefatura': 'Squad Lead',
        'horas': {1: 1, 7: 2}  # Etapas 1 y 7
    },
    {
        'nombre': 'Analista RRLL',
        'cargo': 'Analista Relaciones Laborales',
        'equipo': 'RRLL Central',
        'jefatura': 'Lider Comite Multas',
        'horas': {2: 2, 3: 3, 4: 4, 8: 3}  # Etapas 2, 3, 4, 8
    },
    {
        'nombre': 'ECO (Ejecutivo Clima Organizacional)',
        'cargo': 'ECO',
        'equipo': 'Gestion Personas',
        'jefatura': 'LPM',
        'horas': {3: 2, 5: 3, 7: 2, 8: 2}  # Etapas 3, 5, 7, 8
    },
    {
        'nombre': 'LPM (Lider Personas Mercado)',
        'cargo': 'Lider Personas Mercado',
        'equipo': 'Gestion Personas',
        'jefatura': 'HRBP',
        'horas': {5: 2, 7: 3}  # Etapas 5 y 7
    },
    {
        'nombre': 'Lider Comite de Multas',
        'cargo': 'Lider Comite Multas',
        'equipo': 'RRLL Central',
        'jefatura': 'Gerente RRLL',
        'horas': {6: 4, 10: 3}  # Etapas 6 y 10
    },
    {
        'nombre': 'Legal',
        'cargo': 'Abogado Laboral',
        'equipo': 'Legal',
        'jefatura': 'Gerente Legal',
        'horas': {5: 2, 9: 4}  # Etapas 5 y 9
    },
    {
        'nombre': 'Finanzas',
        'cargo': 'Analista Finanzas',
        'equipo': 'Finanzas',
        'jefatura': 'Controller',
        'horas': {9: 2}  # Etapa 9
    },
]

print('=== LLENANDO HOJA FTEs ===')
print()

# Escribir datos de responsables
fila_inicio = 5  # Fila donde empiezan los datos (despues de headers)

for i, resp in enumerate(responsables_data):
    fila = fila_inicio + i
    print(f'Fila {fila}: {resp["nombre"]}')
    
    # Columnas A-E: Datos del responsable
    ws_ftes.cell(row=fila, column=1, value=f'RUT-{i+1}')  # A: Rut (placeholder)
    ws_ftes.cell(row=fila, column=2, value=resp['nombre'])  # B: Nombre
    ws_ftes.cell(row=fila, column=3, value=resp['cargo'])   # C: Cargo
    ws_ftes.cell(row=fila, column=4, value=resp['equipo'])  # D: Equipo
    ws_ftes.cell(row=fila, column=5, value=resp['jefatura']) # E: Jefatura
    
    # Escribir horas por etapa
    total_horas = 0
    for etapa_num, horas in resp['horas'].items():
        col = etapa_a_columna.get(etapa_num)
        if col:
            ws_ftes.cell(row=fila, column=col, value=horas)
            total_horas += horas
            print(f'  Etapa {etapa_num} -> Col {col} = {horas} hrs')
    
    # Total horas semanales (columna AB = 28)
    ws_ftes.cell(row=fila, column=28, value=total_horas)
    print(f'  Total: {total_horas} hrs')
    print()

print('Guardando archivo...')
wb.save(dst)

# Copiar de vuelta
shutil.copy2(dst, src)
print('Archivo actualizado y copiado a OneDrive!')
