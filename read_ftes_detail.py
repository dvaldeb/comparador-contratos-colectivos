import sys, os
sys.stdout.reconfigure(encoding='utf-8')
import openpyxl
from openpyxl.utils import get_column_letter

wb = openpyxl.load_workbook('temp_v2.xlsx')
ws = wb["FTE\u00b4s"]

print('=== Row 2 (Category headers) ===', flush=True)
for c in range(1, 41):
    v = ws.cell(row=2, column=c).value
    if v:
        print(f'  {get_column_letter(c)}2 = {v}', flush=True)

print('\n=== Row 3 (Stage headers) ===', flush=True)
for c in range(1, 41):
    v = ws.cell(row=3, column=c).value
    if v:
        print(f'  {get_column_letter(c)}3 = {v}', flush=True)

print('\n=== Row 4 (Hr/FTE) ===', flush=True)
for c in range(1, 41):
    v = ws.cell(row=4, column=c).value
    if v:
        print(f'  {get_column_letter(c)}4 = {v}', flush=True)

print('\n=== Row 5 (Ejemplo) ===', flush=True)
for c in range(1, 41):
    v = ws.cell(row=5, column=c).value
    if v is not None:
        print(f'  {get_column_letter(c)}5 = {v}', flush=True)

print('\n=== Merged cells ===', flush=True)
for mc in ws.merged_cells.ranges:
    print(f'  {mc}', flush=True)
