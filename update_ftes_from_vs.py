# -*- coding: utf-8 -*-
from openpyxl import load_workbook
import shutil
import re

# Copiar archivo
src = r'c:\Users\dvaldeb\OneDrive - Walmart Inc\Escritorio\Mapeo FTEs  Atención al Colaborador - Comite de Multas.xlsx'
dst = r'c:\Users\dvaldeb\pupy inteeligente\Comite_Multas_UPDATED.xlsx'
shutil.copy2(src, dst)

wb = load_workbook(dst)
ws_vs = wb.worksheets[1]  # Value Stream -
ws_ftes = wb.worksheets[2]  # FTEs

print('=== LEYENDO VALUE STREAM ACTUALIZADO ===')
print()

# Leer etapas y responsables del Value Stream
etapas_responsables = []
for row in range(2, 15):
    n = ws_vs.cell(row=row, column=1).value
    etapa = ws_vs.cell(row=row, column=3).value
    responsable = ws_vs.cell(row=row, column=6).value
    
    if n is not None and etapa and responsable:
        etapas_responsables.append({
            'num': int(n),
            'etapa': etapa,
            'responsable': responsable
        })
        print(f'Etapa {n}: {responsable}')

print()

# Normalizar roles
def normalizar_rol(rol):
    rol = rol.strip()
    rol = rol.replace(')', '').replace('(', '')
    # Normalizar variantes
    normalizacion = {
        'eco': 'ECO',
        'Eco': 'ECO',
        'ECO': 'ECO',
        'lpt': 'LPT',
        'Lpt': 'LPT',
        'LPT ': 'LPT',
        'LPT': 'LPT',
        'lpm': 'LPM',
        'Lpm': 'LPM',
        'LPM': 'LPM',
        'legal': 'Legal',
        'Legal ': 'Legal',
        'Legal': 'Legal',
        'gt': 'GT',
        'GT': 'GT',
        'adp': 'ADP',
        'ADP': 'ADP',
        'Operaciones LPT': 'Operaciones',
        'MEJORA CONTINUA COMPLIANCE': 'Mejora Continua Compliance',
    }
    return normalizacion.get(rol, rol)

# Extraer roles únicos y mapear a etapas
roles_etapas = {}  # {rol: [etapas donde participa]}

for er in etapas_responsables:
    # Separar roles (pueden estar separados por /, , )
    roles_raw = re.split(r'[/,]', er['responsable'])
    for rol in roles_raw:
        rol = normalizar_rol(rol)
        if rol and len(rol) > 1:
            if rol not in roles_etapas:
                roles_etapas[rol] = []
            if er['num'] not in roles_etapas[rol]:
                roles_etapas[rol].append(er['num'])

print('=== ROLES NORMALIZADOS Y SUS ETAPAS ===')
for rol, etapas in sorted(roles_etapas.items()):
    print(f'  {rol}: Etapas {etapas}')

print()

# Limpiar hoja FTEs (filas 5-15)
print('Limpiando hoja FTEs...')
for row in range(5, 16):
    for col in range(1, 40):
        try:
            ws_ftes.cell(row=row, column=col).value = None
        except:
            pass

# Mapeo de etapa a columna Hr
etapa_col = {1:6, 2:8, 3:10, 4:12, 5:14, 6:16, 7:18, 8:20, 9:22, 10:24}

# Horas estimadas por etapa
horas_por_etapa = {
    1: 2,  # Detección
    2: 3,  # Notificación
    3: 2,  # Registro
    4: 3,  # Categorización
    5: 3,  # Análisis
    6: 4,  # Sesión Comité
    7: 3,  # Plan Acción
    8: 2,  # Seguimiento
    9: 4,  # Resolución
    10: 3, # Consolidación
}

# Info de roles
info_roles = {
    'LPT': {'cargo': 'Líder Personas Tienda', 'equipo': 'Operaciones Tienda', 'jefatura': 'LPM'},
    'GT': {'cargo': 'Gerente de Tienda', 'equipo': 'Operaciones Tienda', 'jefatura': 'Squad Lead'},
    'Operaciones': {'cargo': 'Operaciones Tienda', 'equipo': 'Operaciones', 'jefatura': 'LPM'},
    'ADP': {'cargo': 'Analista de Datos Personas', 'equipo': 'People Analytics', 'jefatura': 'Gerente ADP'},
    'LPM': {'cargo': 'Líder Personas Mercado', 'equipo': 'Gestión Personas', 'jefatura': 'HRBP'},
    'Mejora Continua Compliance': {'cargo': 'Analista Mejora Continua', 'equipo': 'Compliance', 'jefatura': 'Gerente Compliance'},
    'ECO': {'cargo': 'Ejecutivo Clima Organizacional', 'equipo': 'Gestión Personas', 'jefatura': 'LPM'},
    'Legal': {'cargo': 'Abogado Laboral', 'equipo': 'Legal', 'jefatura': 'Gerente Legal'},
}

print('=== LLENANDO HOJA FTEs ===')
print()

fila = 5
for rol, etapas in sorted(roles_etapas.items()):
    if not rol or len(rol) < 2:
        continue
    
    # Obtener info del rol
    info = info_roles.get(rol, {'cargo': rol, 'equipo': 'Por definir', 'jefatura': 'Por definir'})
    
    # Calcular horas
    horas_dict = {}
    for et in etapas:
        if et in horas_por_etapa:
            horas_dict[et] = horas_por_etapa[et]
    
    total_hrs = sum(horas_dict.values())
    
    # Escribir datos
    ws_ftes.cell(row=fila, column=1, value=f'RUT-{fila-4}')
    ws_ftes.cell(row=fila, column=2, value=rol)
    ws_ftes.cell(row=fila, column=3, value=info['cargo'])
    ws_ftes.cell(row=fila, column=4, value=info['equipo'])
    ws_ftes.cell(row=fila, column=5, value=info['jefatura'])
    
    # Escribir horas por etapa
    for et, hrs in horas_dict.items():
        col = etapa_col.get(et)
        if col:
            ws_ftes.cell(row=fila, column=col, value=hrs)
    
    print(f'Fila {fila}: {rol}')
    print(f'         Etapas: {sorted(etapas)} | Total: {total_hrs} hrs')
    
    fila += 1

print()
print('Guardando archivo...')
wb.save(dst)

# Copiar de vuelta
shutil.copy2(dst, src)
print('Archivo actualizado en OneDrive!')

print()
print('=== RESUMEN FINAL ===')
print(f'Total responsables agregados: {fila - 5}')
